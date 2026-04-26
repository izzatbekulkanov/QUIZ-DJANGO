import io
import json
import random
import base64
import uuid
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

import openpyxl
from django.db.models import F, Q, Prefetch
from django.db import transaction

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Count, Avg, Max
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now, make_aware
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.common.context_processors import build_office_helper_session_payload
from apps.question.forms import TestForm, AddQuestionForm
from apps.question.models import Test, Category, Question, Answer, StudentTestAssignment, StudentTest, StudentTestQuestion, HelpResultPlan
from apps.question.utils.access import resolve_assignment_access
from apps.question.utils.schema import ensure_help_result_plan_schema, ensure_student_test_question_proctoring_schema
from apps.question.utils.utils import clean_import_answer_text, clean_import_question_text


TEST_STUDENT_THROUGH = Test.students.through


def build_test_management_summary(test):
    question_summary = Question.objects.filter(test=test).aggregate(
        question_count=Count('id'),
        latest_question_update=Max('updated_at'),
    )
    return {
        'question_count': question_summary['question_count'] or 0,
        'student_count': test.students.count(),
        'assignment_count': test.assignments.count(),
        'latest_question_update': question_summary['latest_question_update'],
    }


def user_can_manage_test_questions(user, test):
    return bool(user.is_superuser or test.created_by_id == user.id)


def get_test_question_usage_count(test_id):
    return StudentTestQuestion.objects.filter(question__test_id=test_id).count()


def get_question_usage_count(question_id):
    return StudentTestQuestion.objects.filter(question_id=question_id).count()


def attach_test_listing_counts(test_items):
    test_ids = [test.id for test in test_items]
    if not test_ids:
        return

    question_counts = {
        item['test_id']: item['total']
        for item in (
            Question.objects.filter(test_id__in=test_ids)
            .values('test_id')
            .annotate(total=Count('id'))
        )
    }
    student_counts = {
        item['test_id']: item['total']
        for item in (
            TEST_STUDENT_THROUGH.objects.filter(test_id__in=test_ids)
            .values('test_id')
            .annotate(total=Count('id'))
        )
    }

    for test in test_items:
        test.question_count = question_counts.get(test.id, 0)
        test.student_count = student_counts.get(test.id, 0)


def build_assignment_overview():
    assignments_queryset = StudentTestAssignment.objects.all()
    return {
        'total_assignments_count': assignments_queryset.count(),
        'active_assignments_count': assignments_queryset.filter(is_active=True).count(),
        'completed_assignments_count': assignments_queryset.filter(status='completed').count(),
        'unique_tests_count': assignments_queryset.order_by().values('test_id').distinct().count(),
        'latest_assignment_created': assignments_queryset.order_by('-created_at').values_list('created_at', flat=True).first(),
    }


def build_assignment_form_context():
    overview = build_assignment_overview()
    overview.update({
        'categories_count': Category.objects.count(),
        'tests_count': Test.objects.count(),
    })
    return overview


def build_assignment_detail_summary(assignment):
    student_tests_queryset = assignment.student_tests.all()
    return {
        'student_test_count': student_tests_queryset.count(),
        'completed_student_count': student_tests_queryset.filter(completed=True).count(),
        'question_pool_count': assignment.test.questions.count(),
        'latest_attempt_time': student_tests_queryset.order_by('-start_time').values_list('start_time', flat=True).first(),
    }


@method_decorator(login_required, name='dispatch')
class QuestionView(View):
    template_name = 'question/views/questions.html'

    def get(self, request):
        category_filter = request.GET.get('category', '')
        search_query = request.GET.get('search', '').strip()

        tests_queryset = Test.objects.select_related('category', 'created_by')
        if category_filter:
            tests_queryset = tests_queryset.filter(category__id=category_filter)
        if search_query:
            tests_queryset = tests_queryset.filter(name__icontains=search_query)

        tests_queryset = tests_queryset.order_by('-updated_at', '-id')

        paginator = Paginator(tests_queryset, 24)
        page_number = request.GET.get('page')
        tests = paginator.get_page(page_number)
        page_tests = list(tests.object_list)
        attach_test_listing_counts(page_tests)
        tests.object_list = page_tests

        categories = Category.objects.order_by('name')
        query_params = request.GET.copy()
        query_params.pop('page', None)
        pagination_query = urlencode(query_params, doseq=True)

        total_tests_count = Test.objects.count()
        filtered_tests_count = paginator.count
        total_questions_count = Question.objects.count()
        active_categories_count = Test.objects.order_by().values('category_id').distinct().count()
        latest_test_update = Test.objects.order_by('-updated_at').values_list('updated_at', flat=True).first()

        context = {
            'tests': tests,
            'categories': categories,
            'filters': {
                'category': category_filter,
                'search': search_query,
            },
            'pagination_query': pagination_query,
            'total_tests_count': total_tests_count,
            'filtered_tests_count': filtered_tests_count,
            'total_questions_count': total_questions_count,
            'active_categories_count': active_categories_count,
            'latest_test_update': latest_test_update,
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class AddQuestionView(View):
    template_name = 'question/views/add-question.html'

    def get(self, request):
        # Fetch categories for the dropdown
        categories = Category.objects.all()

        context = {
            'categories': categories,  # Categories for the dropdown
        }
        return render(request, self.template_name, context)

    def post(self, request):
        # Handle form submission for adding a new test
        category_id = request.POST.get('category')
        name = request.POST.get('name')
        description = request.POST.get('description')

        try:
            category = Category.objects.get(id=category_id)
            Test.objects.create(
                category=category,
                name=name,
                description=description,
                created_by=request.user,
            )
            return JsonResponse({"success": True, "message": "Yangi test muvaffaqiyatli qo'shildi!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Xatolik yuz berdi: {str(e)}"})


@method_decorator(login_required, name='dispatch')
class DeleteTestView(View):
    def post(self, request, test_id):
        if not request.user.is_superuser:
            return JsonResponse({"success": False, "message": "Ruxsat etilmagan amal!"}, status=403)

        test = get_object_or_404(
            Test.objects.annotate(question_count=Count('questions', distinct=True)),
            id=test_id,
        )

        if test.question_count > 0:
            return JsonResponse({
                "success": False,
                "message": "Test ichida savollar mavjudligi sabab o'chirib bo'lmaydi.",
            })

        test.delete()
        return JsonResponse({"success": True, "message": "Test muvaffaqiyatli o'chirildi!"})

        print(f"[DEBUG] POST request received for DeleteTestView with test_id: {test_id}")
        print(f"[DEBUG] User: {request.user}, Is superuser: {request.user.is_superuser}")
        print(f"[DEBUG] CSRF token from header: {request.META.get('HTTP_X_CSRFTOKEN', 'Not provided')}")
        print(f"[DEBUG] Expected CSRF token: {get_token(request)}")

        if request.user.is_superuser:  # Only admins can delete tests
            print(f"[DEBUG] User is superuser, proceeding with test deletion")
            test = get_object_or_404(Test.objects.annotate(question_count=Count('questions')), id=test_id)
            print(f"[DEBUG] Test found: {test}, Question count: {test.question_count}")

            # Check if there are related questions or results
            if test.question_count == 0:
                print(f"[DEBUG] No related questions, deleting test: {test}")
                test.delete()
                print(f"[DEBUG] Test deleted successfully")
                return JsonResponse({"success": True, "message": "Test muvaffaqiyatli o‘chirildi!"})
            else:
                print(f"[DEBUG] Test has {test.question_count} related questions, cannot delete")
                return JsonResponse({
                    "success": False,
                    "message": "Testga bog‘langan savollar yoki natijalar mavjudligi sabab o‘chirib bo‘lmaydi."
                })
        print(f"[DEBUG] User is not superuser, access denied")
        return JsonResponse({"success": False, "message": "Ruxsat etilmagan amal!"})

@method_decorator(login_required, name='dispatch')
class EditTestView(View):
    template_name = 'question/views/edit-question.html'

    def get(self, request, test_id):
        test = get_object_or_404(Test.objects.select_related('category', 'created_by'), id=test_id)
        form = TestForm(instance=test)
        context = {
            'form': form,
            'test': test,
            'summary': build_test_management_summary(test),
            'categories_count': Category.objects.count(),
        }
        return render(request, self.template_name, context)

    def post(self, request, test_id):
        test = get_object_or_404(Test, id=test_id)
        form = TestForm(request.POST, instance=test)
        if form.is_valid():
            form.save()
            return JsonResponse({
                "success": True,
                "message": "Test muvaffaqiyatli yangilandi!",
                "redirect_url": reverse("administrator:question"),
            })
        return JsonResponse({"success": False, "message": "Formada xatolik mavjud!"})


@method_decorator(login_required, name='dispatch')
class AddQuestionTestView(View):
    template_name = 'question/views/question-add-test.html'

    def get(self, request, test_id):
        test = get_object_or_404(Test.objects.select_related('category', 'created_by'), id=test_id)
        form = AddQuestionForm()

        context = {
            'form': form,
            'test': test,
            'summary': build_test_management_summary(test),
            'office_helper_session': build_office_helper_session_payload(request),
        }
        return render(request, self.template_name, context)

    def post(self, request, test_id):
        test = get_object_or_404(Test, id=test_id)
        form = AddQuestionForm(request.POST)

        if form.is_valid():
            question = Question.objects.create(
                test=test,
                text=form.cleaned_data['text'],
            )

            answers = form.cleaned_data['answers']
            is_correct_flags = form.cleaned_data['is_correct_flags']
            for index, answer_text in enumerate(answers):
                Answer.objects.create(
                    question=question,
                    text=answer_text,
                    is_correct=is_correct_flags[index] == 'true',
                )

            return JsonResponse({"success": True, "message": "Savol va barcha javoblar muvaffaqiyatli saqlandi!"})

        return JsonResponse({"success": False, "errors": form.errors})


@method_decorator(login_required, name='dispatch')
class TestQuestionsView(View):
    template_name = 'question/views/test-questions.html'
    default_page_size = 40
    allowed_page_sizes = (20, 40, 80, 120)

    def get(self, request, test_id):
        test = get_object_or_404(Test.objects.select_related('category', 'created_by'), id=test_id)

        search_query = (request.GET.get('search') or '').strip()
        try:
            page_size = int(request.GET.get('page_size') or self.default_page_size)
        except (TypeError, ValueError):
            page_size = self.default_page_size
        if page_size not in self.allowed_page_sizes:
            page_size = self.default_page_size

        questions_queryset = (
            Question.objects
            .filter(test=test)
            .only('id', 'test_id', 'text', 'updated_at')
            .order_by('id')
        )

        if search_query:
            search_filter = Q(text__icontains=search_query)
            if search_query.isdigit():
                search_filter |= Q(id=int(search_query))
            questions_queryset = questions_queryset.filter(search_filter)

        paginator = Paginator(questions_queryset, page_size)
        questions = paginator.get_page(request.GET.get('page'))
        paginated_questions = list(questions.object_list)
        question_ids = [question.id for question in paginated_questions]

        answer_counts = {}
        correct_answer_counts = {}
        usage_counts = {}
        answers_by_question = {}

        if question_ids:
            for item in (
                Answer.objects
                .filter(question_id__in=question_ids)
                .values('question_id')
                .annotate(
                    total=Count('id'),
                    correct_total=Count('id', filter=Q(is_correct=True)),
                )
            ):
                answer_counts[item['question_id']] = item['total']
                correct_answer_counts[item['question_id']] = item['correct_total']

            usage_counts = {
                item['question_id']: item['total']
                for item in (
                    StudentTestQuestion.objects
                    .filter(question_id__in=question_ids)
                    .values('question_id')
                    .annotate(total=Count('id'))
                )
            }

            for answer in (
                Answer.objects
                .filter(question_id__in=question_ids)
                .only('id', 'question_id', 'text', 'is_correct')
                .order_by('id')
            ):
                answers_by_question.setdefault(answer.question_id, []).append(answer)

        for question in paginated_questions:
            question.answer_count = answer_counts.get(question.id, 0)
            question.correct_answer_count = correct_answer_counts.get(question.id, 0)
            question.usage_count = usage_counts.get(question.id, 0)
            question.prefetched_answers = answers_by_question.get(question.id, [])

        questions.object_list = paginated_questions
        question_usage_count = get_test_question_usage_count(test.id)
        summary = build_test_management_summary(test)
        answer_summary = Answer.objects.filter(question__test=test).aggregate(
            total_answers_count=Count('id'),
            total_correct_answer_count=Count('id', filter=Q(is_correct=True)),
        )
        query_params = request.GET.copy()
        query_params.pop('page', None)

        context = {
            'test': test,
            'questions': questions,
            'summary': summary,
            'total_answers_count': answer_summary['total_answers_count'] or 0,
            'total_correct_answer_count': answer_summary['total_correct_answer_count'] or 0,
            'question_usage_count': question_usage_count,
            'can_delete_all_questions': bool(summary['question_count']) and question_usage_count == 0,
            'filtered_question_count': paginator.count,
            'filters': {
                'search': search_query,
                'page_size': str(page_size),
            },
            'pagination_query': urlencode(query_params, doseq=True),
            'question_index_offset': questions.start_index() - 1 if paginator.count else 0,
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class QuestionDetailApiView(View):
    def get(self, request, test_id, question_id):
        question = get_object_or_404(Question, id=question_id, test_id=test_id)

        # Basic ownership check: test creator or superuser can edit.
        if not user_can_manage_test_questions(request.user, question.test):
            return JsonResponse({"success": False, "errors": "Ruxsat yo'q"}, status=403)

        answers = [
            {
                "id": answer.id,
                "text": clean_import_answer_text(answer.text),
                "is_correct": answer.is_correct,
            }
            for answer in question.answers.order_by('id')
        ]
        return JsonResponse({
            "success": True,
            "question": {
                "id": question.id,
                "text": clean_import_question_text(question.text),
                "answers": answers,
            }
        })


@method_decorator(login_required, name='dispatch')
class QuestionUpdateApiView(View):
    def post(self, request, test_id, question_id):
        question = get_object_or_404(Question, id=question_id, test_id=test_id)

        if not user_can_manage_test_questions(request.user, question.test):
            return JsonResponse({"success": False, "errors": "Ruxsat yo'q"}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except Exception:
            return JsonResponse({"success": False, "errors": "JSON formatda xato"})

        text = clean_import_question_text(payload.get('text') or '').strip()
        answers_payload = payload.get('answers') or []

        if not text:
            return JsonResponse({"success": False, "errors": "Savol matni bo'sh bo'lmasligi kerak"})

        if not isinstance(answers_payload, list) or len(answers_payload) < 2:
            return JsonResponse({"success": False, "errors": "Kamida 2 ta variant bo'lishi kerak"})

        cleaned_answers: list[dict] = []
        for a in answers_payload:
            if not isinstance(a, dict):
                continue
            a_text = clean_import_answer_text(a.get('text') or '').strip()
            if not a_text:
                continue
            cleaned_answers.append({
                "id": a.get("id"),
                "text": a_text,
                "is_correct": bool(a.get("is_correct")),
            })

        if len(cleaned_answers) < 2:
            return JsonResponse({"success": False, "errors": "Kamida 2 ta to'ldirilgan variant bo'lishi kerak"})

        if not any(a["is_correct"] for a in cleaned_answers):
            return JsonResponse({"success": False, "errors": "Kamida bitta to'g'ri javob belgilanmashi kerak"})

        with transaction.atomic():
            question.text = text
            question.save(update_fields=['text'])

            existing = {a.id: a for a in question.answers.all()}
            keep_ids: set[int] = set()

            for a in cleaned_answers:
                a_id = a.get("id")
                if a_id:
                    try:
                        a_id_int = int(a_id)
                    except Exception:
                        a_id_int = 0
                else:
                    a_id_int = 0

                if a_id_int and a_id_int in existing:
                    obj = existing[a_id_int]
                    obj.text = a["text"]
                    obj.is_correct = a["is_correct"]
                    obj.save(update_fields=['text', 'is_correct'])
                    keep_ids.add(obj.id)
                else:
                    obj = Answer.objects.create(
                        question=question,
                        text=a["text"],
                        is_correct=a["is_correct"],
                    )
                    keep_ids.add(obj.id)

            # Delete answers removed by user (if any)
            for a_id, obj in existing.items():
                if a_id not in keep_ids:
                    obj.delete()

        return JsonResponse({"success": True, "message": "Savol muvaffaqiyatli yangilandi!"})


@method_decorator(login_required, name='dispatch')
class QuestionDeleteApiView(View):
    def post(self, request, test_id, question_id):
        question = get_object_or_404(Question, id=question_id, test_id=test_id)

        if not user_can_manage_test_questions(request.user, question.test):
            return JsonResponse({"success": False, "errors": "Ruxsat yo'q"}, status=403)

        usage_count = get_question_usage_count(question.id)
        if usage_count:
            return JsonResponse(
                {
                    "success": False,
                    "errors": (
                        "Bu savol asosida talaba test ishlagan. "
                        "Natijalar mavjud bo'lgani uchun savolni o'chirib bo'lmaydi."
                    ),
                    "used_count": usage_count,
                },
                status=409,
            )

        question.delete()
        return JsonResponse({"success": True, "message": "Savol muvaffaqiyatli o'chirildi!"})


@method_decorator(login_required, name='dispatch')
class QuestionsBulkDeleteApiView(View):
    def post(self, request, test_id):
        test = get_object_or_404(Test, id=test_id)

        if not user_can_manage_test_questions(request.user, test):
            return JsonResponse({"success": False, "errors": "Ruxsat yo'q"}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except Exception:
            payload = {}

        if payload.get("confirmation") != "DELETE":
            return JsonResponse(
                {"success": False, "errors": "O'chirish uchun DELETE so'zini aniq kiriting."},
                status=400,
            )

        question_count = Question.objects.filter(test=test).count()
        if not question_count:
            return JsonResponse({"success": False, "errors": "O'chirish uchun savollar topilmadi."}, status=404)

        usage_count = get_test_question_usage_count(test.id)
        if usage_count:
            return JsonResponse(
                {
                    "success": False,
                    "errors": (
                        "Bu test savollari asosida talaba test ishlagan. "
                        "Natijalar mavjud bo'lgani uchun barcha savollarni o'chirib bo'lmaydi."
                    ),
                    "used_count": usage_count,
                },
                status=409,
            )

        with transaction.atomic():
            deleted_count, _ = Question.objects.filter(test=test).delete()

        return JsonResponse(
            {
                "success": True,
                "deleted_questions": question_count,
                "deleted_objects": deleted_count,
                "message": f"{question_count} ta savol muvaffaqiyatli o'chirildi.",
            }
        )

@method_decorator(login_required, name='dispatch')
class AssignTestView(View):
    template_name = 'question/views/assign-test.html'

    def get(self, request):
        # Filtr parametrlari
        category_id = request.GET.get('category', None)
        test_name = request.GET.get('test_name', None)
        start_time = request.GET.get('start_time', None)
        end_time = request.GET.get('end_time', None)

        # Hammasini olish yoki filtr qo‘llash
        assignments = StudentTestAssignment.objects.select_related('category', 'test', 'teacher')

        if category_id:
            assignments = assignments.filter(category_id=category_id)
        if test_name:
            assignments = assignments.filter(test__name__icontains=test_name)
        if start_time:
            assignments = assignments.filter(start_time__gte=start_time)
        if end_time:
            assignments = assignments.filter(end_time__lte=end_time)

        # Pagination
        paginator = Paginator(assignments, 50)  # Har bir sahifada 50 ta element
        page_number = request.GET.get('page', 1)
        paginated_assignments = paginator.get_page(page_number)

        # Kategoriyalarni olish
        categories = Category.objects.all()

        context = {
            'assignments': paginated_assignments,
            'categories': categories,
            'filters': {
                'category': category_id,
                'test_name': test_name,
                'start_time': start_time,
                'end_time': end_time,
            },
        }
        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class EditAssignTestView(View):
    template_name = 'question/views/edit-assign-test.html'

    def get(self, request, assignment_id):
        # Fetch the assignment or return 404
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)
        categories = Category.objects.all()

        # Prepare context with assignment data and categories
        context = {
            'assignment': assignment,
            'categories': categories,
            'filters': {
                'category_id': str(assignment.category.id),
                'test_id': str(assignment.test.id),
                'total_questions': assignment.total_questions,
                'start_time': assignment.start_time.strftime('%Y-%m-%d %H:%M'),
                'end_time': assignment.end_time.strftime('%Y-%m-%d %H:%M'),
                'duration': assignment.duration,
                'max_attempts': assignment.attempts,
            }
        }
        return render(request, self.template_name, context)

    def post(self, request, assignment_id):
        # Fetch the assignment
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        try:
            # Parse form data
            data = request.POST
            category_id = data.get('category_id')
            test_id = data.get('test_id')
            total_questions = int(data.get('total_questions', 0))
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            duration = int(data.get('duration', 0))
            max_attempts = int(data.get('max_attempts', 0))

            # Validate inputs
            if not all([category_id, test_id, total_questions, start_time, end_time, duration, max_attempts]):
                return JsonResponse({'success': False, 'message': 'Barcha maydonlar to‘ldirilishi kerak.'})

            # Validate category and test
            category = get_object_or_404(Category, id=category_id)
            test = get_object_or_404(Test, id=test_id)

            # Validate total questions against test question count
            test_question_count = test.questions.count()  # Assuming Test model has a related 'questions' field
            if total_questions > test_question_count:
                return JsonResponse({
                    'success': False,
                    'message': f"Savollar soni testdagi savollar sonidan ({test_question_count}) oshmasligi kerak."
                })

            # Validate dates
            from datetime import datetime
            start_dt = datetime.strptime(start_time, '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(end_time, '%Y-%m-%d %H:%M')
            now = datetime.now()

            if start_dt >= end_dt:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti tugash vaqtidan keyin bo‘lishi mumkin emas.'
                })
            if end_dt <= now:
                return JsonResponse({
                    'success': False,
                    'message': 'Tugash vaqti hozirgi vaqtdan oldin bo‘lishi mumkin emas.'
                })

            # Update assignment
            assignment.category = category
            assignment.test = test
            assignment.total_questions = total_questions
            assignment.start_time = start_dt
            assignment.end_time = end_dt
            assignment.duration = duration
            assignment.max_attempts = max_attempts
            assignment.save()

            return JsonResponse({
                'success': True,
                'message': 'Topshiriq muvaffaqiyatli yangilandi.'
            })

        except (ValueError, ValidationError) as e:
            return JsonResponse({
                'success': False,
                'message': str(e) if str(e) else 'Noto‘g‘ri ma’lumotlar kiritildi.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': 'Serverda xatolik yuz berdi. Iltimos, qayta urinib ko‘ring.'
            })


@method_decorator(login_required, name='dispatch')
class ViewAssignTestView(View):
    template_name = 'question/views/view-assign-test.html'

    def get(self, request, assignment_id):
        assignment = get_object_or_404(
            StudentTestAssignment.objects.select_related('category', 'test', 'teacher'),
            id=assignment_id,
        )

        group_name = request.GET.get('group_name', '')
        username = request.GET.get('username', '')
        first_name = request.GET.get('first_name', '')
        full_name = request.GET.get('full_name', '')

        student_tests = assignment.student_tests.select_related('student').order_by('-start_time', '-id')
        if group_name:
            student_tests = student_tests.filter(student__group_name__icontains=group_name)
        if username:
            student_tests = student_tests.filter(student__username__icontains=username)
        if first_name:
            student_tests = student_tests.filter(student__first_name__icontains=first_name)
        if full_name:
            student_tests = student_tests.filter(student__full_name__icontains=full_name)

        paginator = Paginator(student_tests, 25)
        page_number = request.GET.get('page')
        paginated_student_tests = paginator.get_page(page_number)

        completed_queryset = student_tests.filter(completed=True)
        score_summary = completed_queryset.aggregate(
            average_score=Avg('score'),
            top_score=Max('score'),
        )
        total_attempts_count = assignment.student_tests.count()
        groups_count = (
            assignment.student_tests.exclude(student__group_name__isnull=True)
            .exclude(student__group_name='')
            .values('student__group_name')
            .distinct()
            .count()
        )
        query_params = request.GET.copy()
        query_params.pop('page', None)

        context = {
            'assignment': assignment,
            'student_tests': paginated_student_tests,
            'summary': build_assignment_detail_summary(assignment),
            'total_attempts_count': total_attempts_count,
            'filtered_attempts_count': paginator.count,
            'filtered_completed_count': completed_queryset.count(),
            'filtered_incomplete_count': student_tests.filter(completed=False).count(),
            'average_score': score_summary['average_score'],
            'top_score': score_summary['top_score'],
            'groups_count': groups_count,
            'pagination_query': urlencode(query_params, doseq=True),
            'filters': {
                'group_name': group_name,
                'username': username,
                'first_name': first_name,
                'full_name': full_name,
            },
        }
        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class ExportAssignTestView(View):
    def get(self, request, assignment_id):
        # Fetch the assignment
        assignment = get_object_or_404(StudentTestAssignment.objects.select_related('test'), id=assignment_id)

        # Apply filters
        group_name = request.GET.get('group_name', '')
        username = request.GET.get('username', '')
        first_name = request.GET.get('first_name', '')
        full_name = request.GET.get('full_name', '')

        student_tests = assignment.student_tests.select_related('student')
        if group_name:
            student_tests = student_tests.filter(student__group_name__icontains=group_name)
        if username:
            student_tests = student_tests.filter(student__username__icontains=username)
        if first_name:
            student_tests = student_tests.filter(student__first_name__icontains=first_name)
        if full_name:
            student_tests = student_tests.filter(student__full_name__icontains=full_name)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Test_{assignment.test.name}_Results"

        # Define headers
        headers = ['#', 'Username', 'Full Name', 'Group Name', 'Score', 'Completed', 'Duration (MM:SS)']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header
            ws[f"{get_column_letter(col_num)}1"].font = openpyxl.styles.Font(bold=True)

        # Populate data
        for row_num, student_test in enumerate(student_tests, 2):
            ws[f"A{row_num}"] = row_num - 1
            ws[f"B{row_num}"] = student_test.student.username
            ws[f"C{row_num}"] = student_test.student.full_name or ''
            ws[f"D{row_num}"] = student_test.student.group_name or ''
            ws[f"E{row_num}"] = student_test.score
            ws[f"F{row_num}"] = 'Yes' if student_test.completed else 'No'
            ws[f"G{row_num}"] = student_test.get_duration_display()

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Test_{assignment.test.name}_Results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        response.write(buffer.getvalue())
        return response

@method_decorator([login_required, require_POST], name='dispatch')
class DeleteAssignmentView(View):
    def post(self, request, assignment_id):
        try:
            assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

            # Agar topshiriqqa bog'liq StudentTest yozuvlari mavjud bo'lsa, o'chirishni taqiqlash
            if assignment.student_tests.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Bu topshiriqni o‘chirish mumkin emas, chunki unga tegishli talaba testlari mavjud.'
                }, status=400)

            assignment.delete()
            return JsonResponse({
                'success': True,
                'message': 'Topshiriq muvaffaqiyatli o‘chirildi!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Xatolik yuz berdi: {str(e)}'
            }, status=500)


@method_decorator(login_required, name='dispatch')
class AddAssignTestView(View):
    template_name = 'question/views/add-assign-test.html'

    def get(self, request):
        categories = Category.objects.all()
        tests = Test.objects.all()
        context = {
            'categories': categories,
            'tests': tests,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        try:
            category_id = request.POST.get('category_id')
            test_id = request.POST.get('test_id')
            total_questions = int(request.POST.get('total_questions', 0))
            start_time_str = request.POST.get('start_time')
            end_time_str = request.POST.get('end_time')
            duration = request.POST.get('duration')
            max_attempts = int(request.POST.get('max_attempts', 3))  # Yangi maydon

            category = get_object_or_404(Category, id=category_id)
            test = get_object_or_404(Test, id=test_id)

            # Savollar sonini tekshirish
            available_questions = test.questions.count()
            if total_questions > available_questions:
                return JsonResponse({
                    'success': False,
                    'message': f"Testda faqat {available_questions} ta savol mavjud."
                }, status=400)

            # Maksimal urinishlar sonini tekshirish
            if max_attempts < 1:
                return JsonResponse({
                    'success': False,
                    'message': "Maksimal urinishlar soni 1 dan kam bo‘lishi mumkin emas."
                }, status=400)

            # Vaqtni parsing qilish
            try:
                start_time = make_aware(datetime.strptime(start_time_str, '%Y-%m-%d %H:%M'))
                end_time = make_aware(datetime.strptime(end_time_str, '%Y-%m-%d %H:%M'))
            except ValueError as e:
                return JsonResponse({
                    'success': False,
                    'message': 'Vaqt formati noto‘g‘ri (YYYY-MM-DD HH:mm).'
                }, status=400)

            # Vaqt validatsiyasi
            if False:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti hozirgi vaqtdan keyin bo‘lishi kerak.'
                }, status=400)
            if start_time >= end_time:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti tugash vaqtidan oldin bo‘lishi kerak.'
                }, status=400)

            # Topshiriq allaqachon mavjudligini tekshirish
            if StudentTestAssignment.objects.filter(
                category=category,
                test=test,
                start_time=start_time,
                end_time=end_time,
            ).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Bunday test topshirig‘i allaqachon mavjud!'
                }, status=400)

            # Topshiriqni yaratish
            assignment = StudentTestAssignment.objects.create(
                teacher=request.user,
                test=test,
                category=category,
                total_questions=total_questions,
                start_time=start_time,
                end_time=end_time,
                duration=duration,
                is_active=True,
                status='pending',
                attempts=max_attempts  # Yangi maydonni saqlash
            )

            return JsonResponse({
                'success': True,
                'message': 'Test topshirig‘i muvaffaqiyatli qo‘shildi!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Xatolik yuz berdi: {str(e)}'
            }, status=500)
@method_decorator(csrf_exempt, name='dispatch')
class ToggleActiveView(View):
    def post(self, request, assignment_id):
        try:
            data = json.loads(request.body)
            is_active = data.get('is_active', None)

            if is_active is None:
                return JsonResponse({"success": False, "message": "Invalid data received."}, status=400)

            assignment = StudentTestAssignment.objects.get(id=assignment_id)
            assignment.is_active = is_active
            assignment.save()

            status = "activated" if is_active else "deactivated"
            return JsonResponse({"success": True, "message": f"The assignment was successfully {status}."})
        except StudentTestAssignment.DoesNotExist:
            return JsonResponse({"success": False, "message": "Assignment not found."}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)


def parse_assignment_datetime(value):
    return make_aware(datetime.strptime(value, '%Y-%m-%d %H:%M'))


@method_decorator(login_required, name='dispatch')
class AssignTestView(View):
    template_name = 'question/views/assign-test.html'

    def get(self, request):
        category_id = request.GET.get('category', '').strip()
        test_name = request.GET.get('test_name', '').strip()
        start_time = request.GET.get('start_time', '').strip()
        end_time = request.GET.get('end_time', '').strip()

        assignments = (
            StudentTestAssignment.objects.select_related('category', 'test', 'teacher')
            .annotate(
                student_test_count=Count('student_tests', distinct=True),
                completed_student_count=Count(
                    'student_tests',
                    filter=Q(student_tests__completed=True),
                    distinct=True,
                ),
            )
            .order_by('-created_at', '-id')
        )

        if category_id:
            assignments = assignments.filter(category_id=category_id)
        if test_name:
            assignments = assignments.filter(test__name__icontains=test_name)
        if start_time:
            assignments = assignments.filter(start_time__gte=start_time)
        if end_time:
            assignments = assignments.filter(end_time__lte=end_time)

        paginator = Paginator(assignments, 24)
        paginated_assignments = paginator.get_page(request.GET.get('page', 1))

        query_params = request.GET.copy()
        query_params.pop('page', None)
        overview = build_assignment_overview()

        context = {
            'assignments': paginated_assignments,
            'categories': Category.objects.order_by('name'),
            'pagination_query': urlencode(query_params, doseq=True),
            'total_assignments_count': overview['total_assignments_count'],
            'filtered_assignments_count': paginator.count,
            'active_assignments_count': overview['active_assignments_count'],
            'completed_assignments_count': overview['completed_assignments_count'],
            'unique_tests_count': overview['unique_tests_count'],
            'latest_assignment_created': overview['latest_assignment_created'],
            'filters': {
                'category': category_id,
                'test_name': test_name,
                'start_time': start_time,
                'end_time': end_time,
            },
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class EditAssignTestView(View):
    template_name = 'question/views/edit-assign-test.html'

    def get(self, request, assignment_id):
        assignment = get_object_or_404(
            StudentTestAssignment.objects.select_related('category', 'test', 'teacher'),
            id=assignment_id,
        )
        overview = build_assignment_form_context()

        context = {
            'assignment': assignment,
            'categories': Category.objects.order_by('name'),
            'summary': build_assignment_detail_summary(assignment),
            'categories_count': overview['categories_count'],
            'tests_count': overview['tests_count'],
            'active_assignments_count': overview['active_assignments_count'],
            'latest_assignment_created': overview['latest_assignment_created'],
            'filters': {
                'category_id': str(assignment.category_id),
                'test_id': str(assignment.test_id),
                'total_questions': assignment.total_questions,
                'start_time': assignment.start_time.strftime('%Y-%m-%d %H:%M'),
                'end_time': assignment.end_time.strftime('%Y-%m-%d %H:%M'),
                'duration': assignment.duration,
                'max_attempts': assignment.attempts,
            },
        }
        return render(request, self.template_name, context)

    def post(self, request, assignment_id):
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        try:
            category_id = request.POST.get('category_id')
            test_id = request.POST.get('test_id')
            total_questions = int(request.POST.get('total_questions', 0))
            duration = int(request.POST.get('duration', 0))
            max_attempts = int(request.POST.get('max_attempts', 0))
            start_time_value = request.POST.get('start_time')
            end_time_value = request.POST.get('end_time')
        except (TypeError, ValueError):
            return JsonResponse({
                'success': False,
                'message': "Noto'g'ri ma'lumotlar kiritildi."
            }, status=400)

        if not all([category_id, test_id, total_questions, duration, max_attempts, start_time_value, end_time_value]):
            return JsonResponse({
                'success': False,
                'message': "Barcha maydonlar to'ldirilishi kerak."
            }, status=400)

        if duration < 1:
            return JsonResponse({
                'success': False,
                'message': "Davomiylik 1 daqiqadan kam bo'lishi mumkin emas."
            }, status=400)

        if max_attempts < 1:
            return JsonResponse({
                'success': False,
                'message': "Maksimal urinishlar soni 1 dan kam bo'lishi mumkin emas."
            }, status=400)

        category = get_object_or_404(Category, id=category_id)
        test = get_object_or_404(Test, id=test_id)
        available_questions = test.questions.count()
        if total_questions > available_questions:
            return JsonResponse({
                'success': False,
                'message': f"Savollar soni testdagi savollar sonidan ({available_questions}) oshmasligi kerak."
            }, status=400)

        try:
            start_time = parse_assignment_datetime(start_time_value)
            end_time = parse_assignment_datetime(end_time_value)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': "Vaqt formati noto'g'ri (YYYY-MM-DD HH:mm)."
            }, status=400)

        if start_time >= end_time:
            return JsonResponse({
                'success': False,
                'message': "Boshlanish vaqti tugash vaqtidan oldin bo'lishi kerak."
            }, status=400)

        duplicate_exists = StudentTestAssignment.objects.filter(
            category=category,
            test=test,
            start_time=start_time,
            end_time=end_time,
        ).exclude(id=assignment.id).exists()
        if duplicate_exists:
            return JsonResponse({
                'success': False,
                'message': "Bunday test topshirig'i allaqachon mavjud!"
            }, status=400)

        assignment.category = category
        assignment.test = test
        assignment.total_questions = total_questions
        assignment.start_time = start_time
        assignment.end_time = end_time
        assignment.duration = duration
        assignment.attempts = max_attempts
        assignment.save()

        return JsonResponse({
            'success': True,
            'message': "Topshiriq muvaffaqiyatli yangilandi.",
            'redirect_url': reverse('administrator:assign-test'),
        })


@method_decorator([login_required, require_POST], name='dispatch')
class DeleteAssignmentView(View):
    def post(self, request, assignment_id):
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        if assignment.student_tests.exists():
            return JsonResponse({
                'success': False,
                'message': "Bu topshiriqni o'chirib bo'lmaydi, chunki unga tegishli talaba testlari mavjud."
            }, status=400)

        assignment.delete()
        return JsonResponse({
            'success': True,
            'message': "Topshiriq muvaffaqiyatli o'chirildi!"
        })


@method_decorator(login_required, name='dispatch')
class AddAssignTestView(View):
    template_name = 'question/views/add-assign-test.html'

    def get(self, request):
        context = build_assignment_form_context()
        context['categories'] = Category.objects.order_by('name')
        return render(request, self.template_name, context)

    def post(self, request):
        try:
            category_id = request.POST.get('category_id')
            test_id = request.POST.get('test_id')
            total_questions = int(request.POST.get('total_questions', 0))
            duration = int(request.POST.get('duration', 0))
            max_attempts = int(request.POST.get('max_attempts', 3))
            start_time_value = request.POST.get('start_time')
            end_time_value = request.POST.get('end_time')
        except (TypeError, ValueError):
            return JsonResponse({
                'success': False,
                'message': "Noto'g'ri ma'lumotlar kiritildi."
            }, status=400)

        if not all([category_id, test_id, total_questions, duration, max_attempts, start_time_value, end_time_value]):
            return JsonResponse({
                'success': False,
                'message': "Barcha maydonlar to'ldirilishi kerak."
            }, status=400)

        if total_questions < 1:
            return JsonResponse({
                'success': False,
                'message': "Savollar soni 1 dan kam bo'lishi mumkin emas."
            }, status=400)

        if duration < 1:
            return JsonResponse({
                'success': False,
                'message': "Davomiylik 1 daqiqadan kam bo'lishi mumkin emas."
            }, status=400)

        if max_attempts < 1:
            return JsonResponse({
                'success': False,
                'message': "Maksimal urinishlar soni 1 dan kam bo'lishi mumkin emas."
            }, status=400)

        category = get_object_or_404(Category, id=category_id)
        test = get_object_or_404(Test, id=test_id)
        available_questions = test.questions.count()
        if total_questions > available_questions:
            return JsonResponse({
                'success': False,
                'message': f"Testda faqat {available_questions} ta savol mavjud."
            }, status=400)

        try:
            start_time = parse_assignment_datetime(start_time_value)
            end_time = parse_assignment_datetime(end_time_value)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': "Vaqt formati noto'g'ri (YYYY-MM-DD HH:mm)."
            }, status=400)

        if start_time >= end_time:
            return JsonResponse({
                'success': False,
                'message': "Boshlanish vaqti tugash vaqtidan oldin bo'lishi kerak."
            }, status=400)

        duplicate_exists = StudentTestAssignment.objects.filter(
            category=category,
            test=test,
            start_time=start_time,
            end_time=end_time,
        ).exists()
        if duplicate_exists:
            return JsonResponse({
                'success': False,
                'message': "Bunday test topshirig'i allaqachon mavjud!"
            }, status=400)

        StudentTestAssignment.objects.create(
            teacher=request.user,
            test=test,
            category=category,
            total_questions=total_questions,
            start_time=start_time,
            end_time=end_time,
            duration=duration,
            is_active=True,
            status='pending',
            attempts=max_attempts,
        )

        return JsonResponse({
            'success': True,
            'message': "Test topshirig'i muvaffaqiyatli qo'shildi!",
            'redirect_url': reverse('administrator:assign-test'),
        })


@method_decorator([login_required, require_POST], name='dispatch')
class ToggleActiveView(View):
    def post(self, request, assignment_id):
        try:
            data = json.loads(request.body or '{}')
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': "Yuborilgan ma'lumot noto'g'ri."
            }, status=400)

        if 'is_active' not in data:
            return JsonResponse({
                'success': False,
                'message': "Faollik holati yuborilmadi."
            }, status=400)

        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)
        assignment.is_active = bool(data['is_active'])
        assignment.save(update_fields=['is_active'])

        return JsonResponse({
            'success': True,
            'message': "Topshiriq holati muvaffaqiyatli yangilandi."
        })


def _build_secure_assignment_redirect(user, route_name, assignment_id):
    from apps.question.utils.access import build_assignment_access_token

    return reverse(route_name, args=[build_assignment_access_token(user.id, assignment_id)])


def _ordered_student_question_answers(student_question):
    answers = list(student_question.question.answers.all())
    shuffler = random.Random(f"{student_question.student_test_id}:{student_question.id}:{student_question.question_id}")
    shuffler.shuffle(answers)
    return answers


def _serialize_student_question(student_question):
    answers = _ordered_student_question_answers(student_question)
    return {
        'id': student_question.id,
        'text': student_question.question.text,
        'image': student_question.question.image.url if student_question.question.image else None,
        'answers': [
            {
                'id': answer.id,
                'text': answer.text,
                'selected': student_question.selected_answer_id == answer.id,
            }
            for answer in answers
        ],
    }


def _student_question_queryset(student_test_id):
    answer_queryset = Answer.objects.only('id', 'question_id', 'text').order_by('id')
    return (
        StudentTestQuestion.objects
        .filter(student_test_id=student_test_id)
        .select_related('question')
        .only(
            'id',
            'student_test_id',
            'question_id',
            'selected_answer_id',
            'order',
            'question__id',
            'question__text',
            'question__image',
        )
        .prefetch_related(Prefetch('question__answers', queryset=answer_queryset))
        .order_by('order')
    )


def _parse_save_answer_request(request):
    content_type = (request.content_type or '').split(';', 1)[0].strip().lower()

    if content_type == 'application/json':
        try:
            payload = json.loads(request.body or '{}')
        except json.JSONDecodeError as exc:
            raise ValidationError("Javob ma'lumotlarini o'qib bo'lmadi.") from exc

        return {
            'student_question_id': payload.get('student_question_id'),
            'answer_id': payload.get('answer_id'),
            'camera_snapshot_data_url': payload.get('camera_snapshot'),
            'camera_snapshot_file': None,
        }

    return {
        'student_question_id': request.POST.get('student_question_id'),
        'answer_id': request.POST.get('answer_id'),
        'camera_snapshot_data_url': '',
        'camera_snapshot_file': request.FILES.get('camera_snapshot'),
    }


def _save_student_answer_snapshot(student_question, snapshot_data_url='', snapshot_file=None):
    if snapshot_file:
        mime_type = (getattr(snapshot_file, 'content_type', '') or '').strip().lower()
        allowed_snapshot_types = {
            'image/jpeg': 'jpg',
            'image/jpg': 'jpg',
            'image/png': 'png',
            'image/webp': 'webp',
        }
        if mime_type and mime_type not in allowed_snapshot_types:
            raise ValidationError("Camera snapshot turi qo'llab-quvvatlanmaydi.")

        extension = Path(getattr(snapshot_file, 'name', '') or '').suffix.lower().lstrip('.')
        if extension == 'jpeg':
            extension = 'jpg'
        if not extension:
            extension = allowed_snapshot_types.get(mime_type, 'jpg')

        if student_question.answer_snapshot:
            student_question.answer_snapshot.delete(save=False)

        file_name = f"question_answer_{student_question.id}_{uuid.uuid4().hex[:10]}.{extension}"
        student_question.answer_snapshot.save(file_name, snapshot_file, save=False)
        return True

    if not snapshot_data_url or not isinstance(snapshot_data_url, str):
        return False

    if not snapshot_data_url.startswith('data:image/'):
        raise ValidationError("Camera snapshot formati noto'g'ri.")

    try:
        header, encoded = snapshot_data_url.split(';base64,', 1)
    except ValueError as exc:
        raise ValidationError("Camera snapshotni o'qib bo'lmadi.") from exc

    mime_type = header.replace('data:', '', 1).strip().lower()
    extension = {
        'image/jpeg': 'jpg',
        'image/jpg': 'jpg',
        'image/png': 'png',
        'image/webp': 'webp',
    }.get(mime_type)
    if not extension:
        raise ValidationError("Camera snapshot turi qo'llab-quvvatlanmaydi.")

    try:
        decoded_file = base64.b64decode(encoded)
    except Exception as exc:
        raise ValidationError("Camera snapshotni dekod qilishda xatolik yuz berdi.") from exc

    if student_question.answer_snapshot:
        student_question.answer_snapshot.delete(save=False)

    file_name = f"question_answer_{student_question.id}_{uuid.uuid4().hex[:10]}.{extension}"
    student_question.answer_snapshot.save(file_name, ContentFile(decoded_file), save=False)
    return True


def _apply_pending_help_result_plan(student_test):
    if not ensure_help_result_plan_schema():
        return None

    plan = (
        HelpResultPlan.objects
        .select_for_update()
        .filter(
            student_id=student_test.student_id,
            assignment_id=student_test.assignment_id,
            status=HelpResultPlan.STATUS_PENDING,
        )
        .order_by('-created_at', '-id')
        .first()
    )
    if not plan:
        return None

    answered_questions = list(
        StudentTestQuestion.objects
        .select_for_update()
        .filter(student_test=student_test, selected_answer__isnull=False)
        .only('id', 'student_test_id', 'question_id', 'selected_answer_id', 'is_correct', 'order')
        .order_by('order', 'id')
    )

    answered_count = len(answered_questions)
    target_correct_count = min(plan.target_correct_answers, answered_count)
    current_correct_count = sum(1 for item in answered_questions if item.is_correct)

    if current_correct_count < target_correct_count:
        needed_correct_count = target_correct_count - current_correct_count
        incorrect_questions = [item for item in answered_questions if not item.is_correct]
        random.shuffle(incorrect_questions)

        correct_answer_by_question = {}
        question_ids = [item.question_id for item in incorrect_questions]
        if question_ids:
            for question_id, answer_id in (
                Answer.objects
                .filter(question_id__in=question_ids, is_correct=True)
                .order_by('question_id', 'id')
                .values_list('question_id', 'id')
            ):
                correct_answer_by_question.setdefault(question_id, answer_id)

        changed_questions = []
        for student_question in incorrect_questions:
            correct_answer_id = correct_answer_by_question.get(student_question.question_id)
            if not correct_answer_id:
                continue

            student_question.selected_answer_id = correct_answer_id
            student_question.is_correct = True
            changed_questions.append(student_question)

            if len(changed_questions) >= needed_correct_count:
                break

        if changed_questions:
            StudentTestQuestion.objects.bulk_update(
                changed_questions,
                ['selected_answer', 'is_correct'],
                batch_size=200,
            )

    plan.status = HelpResultPlan.STATUS_COMPLETED
    plan.save(update_fields=['status', 'updated_at'])
    return plan


@method_decorator(login_required, name='dispatch')
class StartTestDetailView(View):
    template_name = 'test/start_test_detail.html'

    def get(self, request, assignment_id):
        assignment, assignment_token, is_legacy = resolve_assignment_access(request.user, str(assignment_id))
        if is_legacy:
            return HttpResponseRedirect(_build_secure_assignment_redirect(request.user, 'landing:start-test-detail', assignment.id))

        assignment_data = {
            'id': assignment.id,
            'test_name': assignment.test.name,
            'category_name': assignment.category.name,
            'teacher_name': assignment.teacher.full_name,
            'total_questions': assignment.total_questions,
            'start_time': assignment.start_time,
            'end_time': assignment.end_time,
            'duration': assignment.duration,
            'status': assignment.status,
            'is_active': assignment.is_active,
            'created_at': assignment.created_at,
            'token': assignment_token,
            'start_url': reverse('landing:start-test', args=[assignment_token]),
            'unfinished_status_url': reverse('landing:check-unfinished-test', args=[assignment_token]),
        }

        return render(request, self.template_name, {'assignment': assignment_data})

@method_decorator(login_required, name='dispatch')
class StartTestView(View):
    template_name = 'test/start_test.html'

    def get(self, request, assignment_id):
        ensure_student_test_question_proctoring_schema()
        assignment, assignment_token, is_legacy = resolve_assignment_access(request.user, str(assignment_id))
        if is_legacy:
            return HttpResponseRedirect(_build_secure_assignment_redirect(request.user, 'landing:start-test', assignment.id))

        unfinished_test = StudentTest.objects.filter(
            student=request.user,
            assignment=assignment,
            completed=False
        ).first()

        if unfinished_test:
            # Mavjud yakunlanmagan testni qayta ochish
            elapsed_time = (now() - unfinished_test.start_time).seconds
            remaining_time = max(0, assignment.duration * 60 - elapsed_time)  # Sekundlarda

            questions_data = [
                _serialize_student_question(student_question)
                for student_question in _student_question_queryset(unfinished_test.id)
            ]

            context = {
                'assignment': assignment,
                'student_test': unfinished_test,
                'questions': questions_data,
                'remaining_time': remaining_time,  # Qolgan vaqtni yuborish
                'assignment_token': assignment_token,
            }
        else:
            available_question_ids = list(
                Question.objects
                .filter(test_id=assignment.test_id)
                .values_list('id', flat=True)
            )
            if len(available_question_ids) < assignment.total_questions:
                return JsonResponse({'success': False, 'message': 'Yetarli savollar mavjud emas.'}, status=400)

            with transaction.atomic():
                student_test = StudentTest.objects.create(
                    student=request.user,
                    assignment=assignment,
                    start_time=now()
                )
                selected_question_ids = random.sample(available_question_ids, assignment.total_questions)
                StudentTestQuestion.objects.bulk_create(
                    [
                        StudentTestQuestion(
                            student_test=student_test,
                            question_id=question_id,
                            order=index,
                        )
                        for index, question_id in enumerate(selected_question_ids, start=1)
                    ],
                    batch_size=200,
                )

            questions_data = [
                _serialize_student_question(student_question)
                for student_question in _student_question_queryset(student_test.id)
            ]

            context = {
                'assignment': assignment,
                'student_test': student_test,
                'questions': questions_data,
                'remaining_time': assignment.duration * 60,  # To'liq vaqt
                'assignment_token': assignment_token,
            }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class GetRemainingTimeView(View):
    def get(self, request, test_id):
        student_test = get_object_or_404(StudentTest, id=test_id, student=request.user)

        if student_test.completed:
            return JsonResponse({'success': False, 'message': 'Test yakunlangan.'}, status=400)

        assignment = student_test.assignment
        elapsed_time = (now() - student_test.start_time).seconds
        remaining_time = max(0, assignment.duration * 60 - elapsed_time)

        return JsonResponse({'success': True, 'remaining_time': remaining_time})
@method_decorator(login_required, name='dispatch')
class SaveAnswerView(View):
    def post(self, request):
        ensure_student_test_question_proctoring_schema()
        try:
            payload = _parse_save_answer_request(request)
            student_question_id = payload.get('student_question_id')
            answer_id = payload.get('answer_id')
            camera_snapshot = payload.get('camera_snapshot_data_url')
            camera_snapshot_file = payload.get('camera_snapshot_file')

            student_question = get_object_or_404(
                StudentTestQuestion.objects.select_related('student_test', 'question'),
                id=student_question_id,
            )
            if student_question.student_test.student_id != request.user.id:
                return JsonResponse({'success': False, 'message': "Bu savolga javob saqlash uchun ruxsat yo'q."}, status=403)
            if student_question.student_test.completed:
                return JsonResponse({'success': False, 'message': "Yakunlangan testga javob saqlab bo'lmaydi."}, status=400)

            selected_answer = get_object_or_404(Answer, id=answer_id, question_id=student_question.question_id)

            student_question.selected_answer = selected_answer
            student_question.is_correct = selected_answer.is_correct
            student_question.answered_at = timezone.now()
            snapshot_saved = _save_student_answer_snapshot(
                student_question,
                snapshot_data_url=camera_snapshot,
                snapshot_file=camera_snapshot_file,
            )
            update_fields = ['selected_answer', 'is_correct', 'answered_at']
            if snapshot_saved:
                update_fields.append('answer_snapshot')
            student_question.save(update_fields=update_fields)

            return JsonResponse({
                'success': True,
                'message': 'Javob muvaffaqiyatli saqlandi!',
                'answered_at': student_question.answered_at.isoformat() if student_question.answered_at else None,
                'snapshot_url': student_question.answer_snapshot.url if student_question.answer_snapshot else None,
            })
        except ValidationError as exc:
            return JsonResponse({'success': False, 'message': str(exc)}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

@method_decorator(login_required, name='dispatch')
class FinishTestView(View):
    def post(self, request, test_id):
        ensure_help_result_plan_schema()

        with transaction.atomic():
            student_test = get_object_or_404(
                StudentTest.objects.select_for_update(),
                id=test_id,
                student=request.user,
            )

            _apply_pending_help_result_plan(student_test)

            # Testni yakunlash
            correct_answers = student_test.student_questions.filter(is_correct=True).count()
            total_questions = student_test.student_questions.count()
            score = (correct_answers / total_questions * 100) if total_questions > 0 else 0.0

            student_test.completed = True
            student_test.end_time = timezone.now()
            student_test.duration = (student_test.end_time - student_test.start_time).seconds  # Soniyalarda
            student_test.score = score
            student_test.save()

        # Natija sahifasiga yo'naltirish uchun natija URLini qaytarish
        result_url = reverse('landing:view_result', args=[student_test.id])
        return JsonResponse({
            'success': True,
            'message': f'Test tugatildi. Natijangiz: {score:.2f}%',
            'redirect_url': result_url
        })

@method_decorator(login_required, name='dispatch')
class TestResultView(View):
    template_name = 'test/test_result.html'

    def get(self, request, test_id):
        student_test = get_object_or_404(StudentTest, id=test_id, student=request.user)
        correct_answers = student_test.student_questions.filter(is_correct=True).count()
        total_questions = student_test.student_questions.count()

        context = {
            'student_test': student_test,
            'correct_answers': correct_answers,
            'total_questions': total_questions,
        }
        return render(request, self.template_name, context)
