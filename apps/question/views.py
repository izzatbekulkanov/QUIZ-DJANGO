import json
import os
import pickle
import re
import tempfile
import time
import unicodedata
import uuid
from urllib.parse import urlencode

from django.db import models
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Avg, Count, Q, Prefetch
from django.db.models.functions import Coalesce
from django.http import HttpResponseForbidden, JsonResponse, StreamingHttpResponse
from django.middleware.csrf import get_token
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.utils.decorators import method_decorator
from datetime import date, timedelta, datetime
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from apps.account.models import CustomUser
from apps.logs.models import Log
from apps.question.models import StudentTest, Category, StudentTestQuestion, Test, StudentTestAssignment, Question


DEFAULT_USER_PASSWORD = "namdpi451"
USERS_STATS_CACHE_TTL = 120
USERS_GROUPS_CACHE_TTL = 300


@method_decorator(login_required, name='dispatch')
class MainView(View):
    template_name = 'question/views/main.html'

    def get(self, request):
        # Statistika ma'lumotlarini hisoblash
        total_categories = Category.objects.count()
        total_tests = Test.objects.count()
        total_questions = Question.objects.count()
        total_users = CustomUser.objects.count()
        
        # So'nggi 5 ta testni olish
        recent_tests = Test.objects.all().order_by('-id')[:5]
        
        context = {
            'total_categories': total_categories,
            'total_tests': total_tests,
            'total_questions': total_questions,
            'total_users': total_users,
            'recent_tests': recent_tests,
        }
        
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class ProfileView(View):
    template_name = 'question/views/profile.html'

    def get(self, request):
        user = request.user

        created_tests = (
            Test.objects.filter(created_by=user)
            .select_related('category')
            .order_by('-created_at')
        )
        assignments = (
            StudentTestAssignment.objects.filter(teacher=user)
            .select_related('test', 'category')
            .order_by('-created_at')
        )
        completed_results = (
            StudentTest.objects.filter(assignment__teacher=user, completed=True)
            .select_related('student', 'assignment__test', 'assignment__category')
            .order_by('-end_time')
        )
        recent_logs = Log.objects.filter(user=user).order_by('-timestamp')[:8]

        average_score = completed_results.aggregate(avg_score=Avg('score'))['avg_score'] or 0
        role_labels = []
        if user.is_superuser:
            role_labels.append("Superadmin")
        if user.is_staff and not user.is_superuser:
            role_labels.append("Administrator")
        if getattr(user, "is_teacher", False):
            role_labels.append("O'qituvchi")
        if getattr(user, "is_student", False):
            role_labels.append("Talaba")
        if not role_labels:
            role_labels.append("Foydalanuvchi")

        context = {
            'profile_user': user,
            'role_labels': role_labels,
            'stats': {
                'created_tests': created_tests.count(),
                'active_assignments': assignments.filter(is_active=True).count(),
                'completed_results': completed_results.count(),
                'activity_logs': Log.objects.filter(user=user).count(),
                'managed_students': (
                    StudentTest.objects.filter(assignment__teacher=user)
                    .values('student_id')
                    .distinct()
                    .count()
                ),
                'question_bank_size': Question.objects.filter(test__created_by=user).count(),
                'average_score': round(average_score, 1),
            },
            'recent_tests': created_tests[:5],
            'recent_assignments': assignments[:5],
            'recent_results': completed_results[:5],
            'recent_logs': recent_logs,
        }
        return render(request, self.template_name, context)


@method_decorator(csrf_exempt, name='dispatch')
class ToggleAuthView(View):
    def post(self, request, id):
        try:
            user = CustomUser.objects.get(id=id)
            data = json.loads(request.body)
            user.auth_is_id = data.get('auth_is_id', False)
            user.save()
            return JsonResponse({"success": True, "message": "Autentifikatsiya sozlamasi yangilandi"})
        except CustomUser.DoesNotExist:
            return JsonResponse({"success": False, "message": "Foydalanuvchi topilmadi"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GrantAllAuthView(View):
    """Barcha foydalanuvchilarga login ruxsatini berish"""
    def post(self, request):
        try:
            updated_count = CustomUser.objects.all().update(auth_is_id=True)
            return JsonResponse({
                "success": True,
                "message": f"{updated_count} ta foydalanuvchiga ruxsat berildi",
                "count": updated_count
            })
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class RevokeAllAuthView(View):
    """Barcha foydalanuvchilardan login ruxsatini olib tashlash"""
    def post(self, request):
        try:
            updated_count = CustomUser.objects.all().update(auth_is_id=False)
            return JsonResponse({
                "success": True,
                "message": f"{updated_count} ta foydalanuvchidan ruxsati olib tashlandi",
                "count": updated_count
            })
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
@method_decorator(login_required, name='dispatch')
class UsersView(View):
    template_name = 'question/views/users.html'
    page_size = 30
    list_fields = (
        'id',
        'username',
        'first_name',
        'second_name',
        'full_name',
        'group_name',
        'is_student',
        'is_teacher',
        'level_name',
        'staff_position',
        'date_of_birth',
        'auth_is_id',
        'profile_picture',
    )

    def _base_queryset(self):
        return CustomUser.objects.only(*self.list_fields)

    def _get_group_names(self):
        cache_key = "administrator_users_group_names"
        cached_groups = cache.get(cache_key)
        if cached_groups is not None:
            return cached_groups

        groups = self._get_group_names()
        cache.set(cache_key, groups, USERS_GROUPS_CACHE_TTL)
        return groups

    def _get_summary_stats(self):
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)
        month_start = today.replace(day=1)
        cache_key = f"administrator_users_stats:{today.isoformat()}"
        cached_stats = cache.get(cache_key)
        if cached_stats is not None:
            return cached_stats

        stats = CustomUser.objects.aggregate(
            total_users_count=Count('id'),
            student_users_count=Count('id', filter=Q(is_student=True)),
            teacher_users_count=Count('id', filter=Q(is_teacher=True)),
            auth_enabled_count=Count('id', filter=Q(auth_is_id=True)),
            today_users_count=Count('id', filter=Q(date_joined__date=today)),
            yesterday_users_count=Count('id', filter=Q(date_joined__date=yesterday)),
            month_users_count=Count(
                'id',
                filter=Q(date_joined__date__gte=month_start, date_joined__date__lte=today),
            ),
        )
        cache.set(cache_key, stats, USERS_STATS_CACHE_TTL)
        return stats

    def get(self, request):
        search_query = (request.GET.get('search') or '').strip()
        filter_type = (request.GET.get('filter_type') or '').strip()
        group_name = (request.GET.get('group_name') or request.GET.get('group_id') or '').strip()

        users_queryset = self._base_queryset()

        # Qidiruv (to‘liq, username ham kiritilgan)
        if search_query:
            users_queryset = users_queryset.filter(
                Q(full_name__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(second_name__icontains=search_query) |
                Q(third_name__icontains=search_query) |
                Q(username__icontains=search_query) |  # ✅ Username bo‘yicha qidiruv
                Q(phone_number__icontains=search_query) |
                Q(address__icontains=search_query) |
                Q(group_name__icontains=search_query)
            )

        # Filtrlash
        if filter_type == 'student':
            users_queryset = users_queryset.filter(is_student=True)
        elif filter_type == 'teacher':
            users_queryset = users_queryset.filter(is_teacher=True)
        elif filter_type == 'other':
            users_queryset = users_queryset.filter(is_student=False, is_teacher=False)
        elif filter_type == 'group' and group_name:
            users_queryset = users_queryset.filter(group_name=group_name)

        # Saralash
        users_queryset = users_queryset.order_by('group_name', 'second_name', 'first_name', 'username')

        # Guruhlar ro‘yxati
        groups = list(
            CustomUser.objects.exclude(group_name__isnull=True)
            .exclude(group_name__exact='')
            .values_list('group_name', flat=True)
            .distinct()
            .order_by('group_name')
        )

        # Pagination
        paginator = Paginator(users_queryset, self.page_size)
        page_number = request.GET.get('page')
        users = paginator.get_page(page_number)

        # Statistik ma'lumotlar
        stats = self._get_summary_stats()

        query_params = request.GET.copy()
        query_params.pop('page', None)

        context = {
            'users': users,
            'search_query': search_query,
            'filter_type': filter_type,
            'group_name': group_name,
            'groups': groups,
            'total_users_count': stats['total_users_count'],
            'student_users_count': stats['student_users_count'],
            'teacher_users_count': stats['teacher_users_count'],
            'auth_enabled_count': stats['auth_enabled_count'],
            'today_users_count': stats['today_users_count'],
            'yesterday_users_count': stats['yesterday_users_count'],
            'month_users_count': stats['month_users_count'],
            'pagination_query': query_params.urlencode(),
        }
        return render(request, self.template_name, context)

    def delete(self, request, id=None):
        try:
            user = CustomUser.objects.get(id=id)
            # Bog'liqliklarni tekshirish
            created_tests = Test.objects.filter(created_by=user).count()
            test_assignments = StudentTestAssignment.objects.filter(teacher=user).count()
            student_tests = StudentTest.objects.filter(student=user).count()
            user_logs = Log.objects.filter(user=user).count()

            if created_tests > 0 or test_assignments > 0 or student_tests > 0:
                error_message = "Bu foydalanuvchini o'chirib bo'lmaydi: "
                if created_tests > 0:
                    error_message += f"{created_tests} ta test yaratgan; "
                if test_assignments > 0:
                    error_message += f"{test_assignments} ta test topshirig'i bergan; "
                if student_tests > 0:
                    error_message += f"{student_tests} ta test sinovida ishtirok etgan."
                if user_logs > 0:
                    error_message += f"{user_logs} ta log yozuvi mavjud."
                return JsonResponse({
                    "success": False,
                    "message": error_message
                }, status=400)

            user.delete()
            return JsonResponse({
                "success": True,
                "message": "Foydalanuvchi muvaffaqiyatli o'chirildi!"
            })

        except CustomUser.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": "Foydalanuvchi topilmadi!"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False,
                "message": f"Xatolik yuz berdi: {str(e)}"
            }, status=500)


@method_decorator(login_required, name='dispatch')
class AddUserView(View):
    template_name = 'question/views/add-user.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        # Forma ma'lumotlarini olish
        first_name = request.POST.get('first_name')
        second_name = request.POST.get('second_name')
        third_name = request.POST.get('third_name')
        phone_number = request.POST.get('phone_number')
        address = request.POST.get('address')
        contact_email = request.POST.get('contact_email')
        date_of_birth = request.POST.get('date_of_birth')
        gender = request.POST.get('gender')
        nationality = request.POST.get('nationality')
        citizenship = request.POST.get('citizenship')
        bio = request.POST.get('bio')
        emergency_contact = request.POST.get('emergency_contact')
        is_teacher = request.POST.get('is_teacher') == 'on'
        is_student = request.POST.get('is_student') == 'on'
        profile_picture = request.FILES.get('profile_picture')
        username = request.POST.get('username')
        job_title = request.POST.get('job_title')
        company_name = request.POST.get('company_name')
        department = request.POST.get('department')
        academic_degree = request.POST.get('academic_degree')
        academic_rank = request.POST.get('academic_rank')
        staff_position = request.POST.get('staff_position')
        student_id_number = request.POST.get('student_id_number')
        group_name = request.POST.get('group_name')
        specialty = request.POST.get('specialty')
        specialty_name = request.POST.get('specialty_name')
        education_level = request.POST.get('education_level')
        education_type = request.POST.get('education_type')
        education_form_name = request.POST.get('education_form_name')
        payment_form = request.POST.get('payment_form')
        education_year = request.POST.get('education_year')
        department_name = request.POST.get('department_name')
        level_name = request.POST.get('level_name')

        # Debug uchun barcha ma'lumotlarni print qilish
        print("=== POST Ma'lumotlari ===")
        print(f"first_name: {first_name}")
        print(f"second_name: {second_name}")
        print(f"third_name: {third_name}")
        print(f"phone_number: {phone_number}")
        print(f"address: {address}")
        print(f"contact_email: {contact_email}")
        print(f"date_of_birth: {date_of_birth}")
        print(f"gender: {gender}")
        print(f"nationality: {nationality}")
        print(f"citizenship: {citizenship}")
        print(f"bio: {bio}")
        print(f"emergency_contact: {emergency_contact}")
        print(f"is_teacher: {is_teacher}")
        print(f"is_student: {is_student}")
        print(f"profile_picture: {profile_picture}")
        print(f"username: {username}")
        print(f"job_title: {job_title}")
        print(f"company_name: {company_name}")
        print(f"department: {department}")
        print(f"academic_degree: {academic_degree}")
        print(f"academic_rank: {academic_rank}")
        print(f"staff_position: {staff_position}")
        print(f"student_id_number: {student_id_number}")
        print(f"group_name: {group_name}")
        print(f"specialty: {specialty}")
        print(f"specialty_name: {specialty_name}")
        print(f"education_level: {education_level}")
        print(f"education_type: {education_type}")
        print(f"education_form_name: {education_form_name}")
        print(f"payment_form: {payment_form}")
        print(f"education_year: {education_year}")
        print(f"department_name: {department_name}")
        print(f"level_name: {level_name}")
        print("====================")

        default_password = DEFAULT_USER_PASSWORD

        # Validatsiya xatoliklari
        errors = {}

        # Asosiy ma'lumotlar validatsiyasi
        if not first_name:
            errors['first_name'] = "Ismni kiritish majburiy."
        if not second_name:
            errors['second_name'] = "Familiyani kiritish majburiy."
        if not username:
            errors['username'] = "Foydalanuvchi nomini kiritish majburiy."
        if not phone_number:
            errors['phone_number'] = "Telefon raqamni kiritish majburiy."
        if not address:
            errors['address'] = "Manzilni kiritish majburiy."
        if not date_of_birth:
            errors['date_of_birth'] = "Tug'ilgan sanani kiritish majburiy."
        if not gender:
            errors['gender'] = "Jinsni tanlash majburiy."
        if not nationality:
            errors['nationality'] = "Millatni kiritish majburiy."
        if not citizenship:
            errors['citizenship'] = "Fuqarolikni kiritish majburiy."

        # O'qituvchi ma'lumotlari validatsiyasi
        if is_teacher:
            if not job_title:
                errors['job_title'] = "Lavozimni kiritish majburiy (o'qituvchi uchun)."
            if not department:
                errors['department'] = "Fakultet/kafedra nomini kiritish majburiy (o'qituvchi uchun)."

        # Talaba ma'lumotlari validatsiyasi
        if is_student:
            if not student_id_number:
                errors['student_id_number'] = "Talaba ID raqamini kiritish majburiy."
            if not group_name:
                errors['group_name'] = "Guruh nomini kiritish majburiy."
            if not specialty:
                errors['specialty'] = "Mutaxassislik kodini kiritish majburiy."
            if not specialty_name:
                errors['specialty_name'] = "Mutaxassislik nomini kiritish majburiy."
            if not education_form_name:
                errors['education_form_name'] = "Ta'lim shaklini kiritish majburiy."
            if not department_name:
                errors['department_name'] = "Fakultet nomini kiritish majburiy."
            if not level_name:
                errors['level_name'] = "Kursni kiritish majburiy."

        if errors:
            print("=== Validatsiya Xatoliklari ===")
            print(errors)
            print("========================")
            return JsonResponse({"success": False, "errors": errors}, status=400)

        # Foydalanuvchini ma'lumotlar bazasiga saqlash
        try:
            user = CustomUser(
                username=username,
                first_name=first_name,
                second_name=second_name,
                third_name=third_name,
                phone_number=phone_number,
                address=address,
                contact_email=contact_email,
                date_of_birth=date_of_birth,
                gender=gender,
                nationality=nationality,
                citizenship=citizenship,
                bio=bio,
                emergency_contact=emergency_contact,
                is_teacher=is_teacher,
                is_student=is_student,
                profile_picture=profile_picture,
                job_title=job_title if is_teacher else None,
                company_name=company_name if is_teacher else None,
                department=department if is_teacher else None,
                academic_degree=academic_degree if is_teacher else None,
                academic_rank=academic_rank if is_teacher else None,
                staff_position=staff_position if is_teacher else None,
                student_id_number=student_id_number if is_student else None,
                group_name=group_name if is_student else None,
                specialty=specialty if is_student else None,
                specialty_name=specialty_name if is_student else None,
                education_level=education_level if is_student else None,
                education_type=education_type if is_student else None,
                education_form_name=education_form_name if is_student else None,
                payment_form=payment_form if is_student else None,
                education_year=education_year if is_student else None,
                department_name=department_name if is_student else None,
                level_name=level_name if is_student else None
            )
            user.set_password(default_password)
            user.save()
            print("=== Foydalanuvchi Muvaffaqiyatli Saqlandi ===")
            print(f"Username: {user.username}")
            print(f"ID: {user.id}")
            print("=================================")
            return JsonResponse({"success": True, "message": "Foydalanuvchi muvaffaqiyatli saqlandi!"}, status=200)
        except ValidationError as e:
            print("=== ValidationError ===")
            print(e.message_dict)
            print("==================")
            return JsonResponse({"success": False, "message": e.message_dict}, status=400)
        except Exception as e:
            print("=== Umumiy Xatolik ===")
            print(str(e))
            print("==================")
            return JsonResponse({"success": False, "message": f"Xatolik yuz berdi: {str(e)}"}, status=500)


@method_decorator(login_required, name='dispatch')
class EditUserView(View):
    template_name = 'question/views/edit-user.html'

    def get(self, request, id):
        # Foydalanuvchini ID bo'yicha topish
        user = get_object_or_404(CustomUser, id=id)
        context = {
            'user': user
        }
        return render(request, self.template_name, context)

    def post(self, request, id):
        # Foydalanuvchini ID bo'yicha topish
        user = get_object_or_404(CustomUser, id=id)

        # Formadan ma'lumotlarni olish
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.first_name = request.POST.get('first_name')
        user.second_name = request.POST.get('second_name')
        user.full_name = request.POST.get('full_name')
        user.phone_number = request.POST.get('phone_number')
        user.address = request.POST.get('address')
        dob_raw = (request.POST.get('date_of_birth') or '').strip()
        # Some clients can send empty values as fancy quotes; normalize to empty.
        dob_raw = (
            dob_raw.strip("\"'")
            .replace("\u201c", "")
            .replace("\u201d", "")
            .replace("вЂњ", "")
            .replace("вЂќ", "")
            .strip()
        )
        if not dob_raw:
            user.date_of_birth = None
        else:
            parsed = None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    parsed = datetime.strptime(dob_raw, fmt).date()
                    break
                except ValueError:
                    continue

            if parsed is None:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Tug'ilgan sana noto'g'ri formatda. YYYY-MM-DD ko'rinishida bo'lishi kerak.",
                    },
                    status=400,
                )

            user.date_of_birth = parsed
        user.gender = request.POST.get('gender')
        user.profile_picture = request.FILES.get('profile_picture', user.profile_picture)
        user.nationality = request.POST.get('nationality')
        user.bio = request.POST.get('bio')
        user.is_student = request.POST.get('is_student') == 'on'
        user.is_teacher = request.POST.get('is_teacher') == 'on'
        user.emergency_contact = request.POST.get('emergency_contact')
        user.job_title = request.POST.get('job_title')
        user.company_name = request.POST.get('company_name')

        # Parolni yangilash
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        if new_password:
            if new_password == confirm_password:
                user.set_password(new_password)
            else:
                return JsonResponse({"success": False, "message": "Parollar mos kelmadi!"}, status=400)

        # Foydalanuvchini saqlash
        try:
            user.save()
            return JsonResponse({"success": True, "message": "Foydalanuvchi muvaffaqiyatli yangilandi!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Xatolik yuz berdi: {str(e)}"}, status=500)


@method_decorator(login_required, name='dispatch')
class ResultsView(View):
    template_name = 'question/views/results.html'
    valid_status_filters = {'all', 'completed', 'incomplete'}
    valid_time_filters = {'today', 'yesterday', 'week', 'month', 'all'}

    def get(self, request):
        category_id = request.GET.get('category', '').strip()
        test_name = request.GET.get('test', '').strip()
        username = request.GET.get('username', '').strip()
        full_name = request.GET.get('full_name', '').strip()
        status_filter = (request.GET.get('status') or 'all').strip().lower() or 'all'
        time_filter = (request.GET.get('time_filter') or 'today').strip().lower() or 'today'

        if status_filter not in self.valid_status_filters:
            status_filter = 'all'
        if time_filter not in self.valid_time_filters:
            time_filter = 'today'

        student_tests = (
            StudentTest.objects.select_related(
                'student',
                'assignment__teacher',
                'assignment__test',
                'assignment__test__created_by',
                'assignment__category',
            )
            .annotate(activity_time=Coalesce('end_time', 'start_time'))
            .order_by('-activity_time', '-id')
        )

        if category_id:
            student_tests = student_tests.filter(assignment__category_id=category_id)
        if test_name:
            student_tests = student_tests.filter(assignment__test__name__icontains=test_name)
        if username:
            student_tests = student_tests.filter(student__username__icontains=username)
        if full_name:
            name_query = Q(student__full_name__icontains=full_name)
            for term in full_name.split():
                name_query |= (
                    Q(student__first_name__icontains=term)
                    | Q(student__second_name__icontains=term)
                    | Q(student__third_name__icontains=term)
                )
            student_tests = student_tests.filter(name_query)

        if status_filter == 'completed':
            student_tests = student_tests.filter(completed=True)
        elif status_filter == 'incomplete':
            student_tests = student_tests.filter(completed=False)

        today = timezone.localdate()
        if time_filter == 'today':
            student_tests = student_tests.filter(activity_time__date=today)
        elif time_filter == 'yesterday':
            student_tests = student_tests.filter(activity_time__date=today - timedelta(days=1))
        elif time_filter == 'week':
            week_start = today - timedelta(days=today.weekday())
            student_tests = student_tests.filter(activity_time__date__gte=week_start, activity_time__date__lte=today)
        elif time_filter == 'month':
            month_start = today.replace(day=1)
            student_tests = student_tests.filter(activity_time__date__gte=month_start, activity_time__date__lte=today)

        categories = Category.objects.all().order_by('name')
        paginator = Paginator(student_tests, 25)
        page_number = request.GET.get('page')
        paginated_student_tests = paginator.get_page(page_number)
        completed_queryset = student_tests.filter(completed=True)
        query_params = request.GET.copy()
        query_params.pop('page', None)

        context = {
            'student_tests': paginated_student_tests,
            'categories': categories,
            'total_results_count': StudentTest.objects.count(),
            'filtered_results_count': paginator.count,
            'completed_results_count': student_tests.filter(completed=True).count(),
            'incomplete_results_count': student_tests.filter(completed=False).count(),
            'average_score': completed_queryset.aggregate(avg_score=Avg('score'))['avg_score'],
            'pagination_query': query_params.urlencode(),
            'filters': {
                'category': category_id,
                'test': test_name,
                'username': username,
                'full_name': full_name,
                'status': status_filter,
                'time_filter': time_filter,
            },
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class DeleteStudentResultView(View):
    def post(self, request, test_id):
        student_test = get_object_or_404(
            StudentTest.objects.select_related('assignment__teacher', 'assignment__test'),
            id=test_id,
        )

        allowed = (
            request.user.is_superuser
            or student_test.assignment.teacher_id == request.user.id
            or student_test.assignment.test.created_by_id == request.user.id
        )
        if not allowed:
            return JsonResponse({"success": False, "message": "Ruxsat yo'q"}, status=403)

        student_label = student_test.student.username
        test_label = student_test.assignment.test.name
        student_test.delete()

        return JsonResponse(
            {
                "success": True,
                "message": f"{student_label} uchun '{test_label}' natijasi o'chirildi. Endi qayta urinish mumkin.",
            }
        )


@method_decorator(login_required, name='dispatch')
class ViewTestDetailsView(View):
    template_name = 'question/views/test_details.html'

    def get(self, request, test_id):
        from apps.question.utils.schema import ensure_student_test_question_proctoring_schema

        ensure_student_test_question_proctoring_schema()
        test = get_object_or_404(
            StudentTest.objects.select_related(
                'student',
                'assignment__teacher',
                'assignment__test',
                'assignment__test__created_by',
                'assignment__category',
            ).prefetch_related(
                Prefetch(
                    'student_questions',
                    queryset=StudentTestQuestion.objects.select_related(
                        'question', 'selected_answer'
                    ).prefetch_related(
                        'question__answers'
                    ).order_by('order')
                )
            ),
            id=test_id
        )

        allowed = (
            request.user.is_superuser
            or test.assignment.teacher_id == request.user.id
            or test.assignment.test.created_by_id == request.user.id
        )
        if not allowed:
            return HttpResponseForbidden("Bu natijani ko'rish uchun ruxsat yo'q.")

        question_rows = list(test.student_questions.all())
        total_questions = len(question_rows)
        answered_count = sum(1 for question in question_rows if question.selected_answer_id)
        correct_answers = sum(1 for question in question_rows if question.selected_answer_id and question.is_correct)
        incorrect_answers = answered_count - correct_answers
        unanswered_count = total_questions - answered_count
        progress_percent = round((answered_count / total_questions) * 100, 1) if total_questions else 0
        accuracy_percent = round((correct_answers / answered_count) * 100, 1) if answered_count else 0

        context = {
            'test': test,
            'question_rows': question_rows,
            'total_questions': total_questions,
            'answered_count': answered_count,
            'correct_answers': correct_answers,
            'incorrect_answers': incorrect_answers,
            'unanswered_count': unanswered_count,
            'progress_percent': progress_percent,
            'accuracy_percent': accuracy_percent,
        }

        return render(request, self.template_name, context)


def assign_students_to_test(request, test_id):
    test = get_object_or_404(Test.objects.select_related('category', 'created_by'), id=test_id)
    students = CustomUser.objects.filter(is_student=True).exclude(assigned_tests=test).order_by('group_name', 'second_name', 'first_name', 'username')
    available_students_total = students.count()

    group_filter = request.GET.get('group', '')
    search_query = request.GET.get('search', '')
    if group_filter:
        students = students.filter(group_name=group_filter)
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(second_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(student_id_number__icontains=search_query)
        )

    groups_qs = (
        CustomUser.objects.filter(is_student=True, group_name__isnull=False)
        .exclude(group_name="")
        .order_by('group_name')
    )
    groups = list(groups_qs.values_list('group_name', flat=True).distinct())
    filtered_students_count = students.count()

    paginator = Paginator(students, 40)
    page_number = request.GET.get('page')
    students_page = paginator.get_page(page_number)

    assigned_students = test.students.order_by('group_name', 'second_name', 'first_name', 'username')
    query_params = request.GET.copy()
    query_params.pop('page', None)

    if request.method == 'POST':
        query_params = {}
        if group_filter:
            query_params['group'] = group_filter
        if search_query:
            query_params['search'] = search_query
        if page_number:
            query_params['page'] = page_number

        redirect_url = reverse('administrator:assign-students-to-test', kwargs={'test_id': test_id})
        if query_params:
            redirect_url += '?' + urlencode(query_params)

        if 'remove_student' in request.POST:
            student_id = request.POST.get('remove_student')
            test.students.remove(student_id)
            query_params['alert'] = 'remove_success'
            redirect_url = reverse('administrator:assign-students-to-test', kwargs={'test_id': test_id}) + '?' + urlencode(
                query_params)
            return redirect(redirect_url)
        else:
            selected_student_ids = request.POST.getlist('students')
            if selected_student_ids:
                test.students.add(*selected_student_ids)
                query_params['alert'] = 'add_success'
                redirect_url = reverse('administrator:assign-students-to-test', kwargs={'test_id': test_id}) + '?' + urlencode(
                    query_params)
                return redirect(redirect_url)

        return redirect(redirect_url)

    alert = None
    if request.GET.get('alert') == 'add_success':
        alert = {'type': 'success', 'message': 'Talabalar muvaffaqiyatli biriktirildi!'}
    elif request.GET.get('alert') == 'remove_success':
        alert = {'type': 'success', 'message': 'Talaba testdan olib tashlandi!'}

    return render(request, 'question/views/assign_students.html', {
        'test': test,
        'students': students_page,
        'assigned_students': assigned_students,
        'groups': groups,
        'group_filter': group_filter,
        'search_query': search_query,
        'alert': alert,
        'pagination_query': query_params.urlencode(),
        'assigned_students_count': assigned_students.count(),
        'available_students_total': available_students_total,
        'filtered_students_count': filtered_students_count,
        'groups_count': len(groups),
        'question_count': test.questions.count(),
    })

EXCEL_IMPORT_COLUMN_ALIASES = {
    "username": {
        "username",
        "talabaid",
        "talabaidraqami",
        "studentid",
        "studentidnumber",
        "studentidraqami",
        "tr",
        "id",
    },
    "full_name": {
        "oquvchiningfish",
        "oquvchiningfio",
        "fish",
        "fio",
        "ismfamiliyasiotasiningismi",
        "ismfamiliyaotasiningismi",
    },
    "group_name": {
        "akademikguruh",
        "akademikguruhi",
        "guruh",
        "group",
        "groupname",
    },
}


def _normalize_excel_header(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = normalized.casefold()
    return re.sub(r"[^a-z0-9]+", "", normalized)


def _clean_excel_value(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def _coerce_excel_username(value):
    text = _clean_excel_value(value)
    if not text:
        return ""

    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return text


def _split_student_full_name(full_name):
    parts = [part for part in full_name.split() if part]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[1], parts[0], ""
    return parts[1], parts[0], " ".join(parts[2:])


def _resolve_excel_columns(columns):
    normalized_columns = {
        _normalize_excel_header(column): column
        for column in columns
    }
    resolved = {}
    for field_name, aliases in EXCEL_IMPORT_COLUMN_ALIASES.items():
        for alias in aliases:
            original_column = normalized_columns.get(alias)
            if original_column:
                resolved[field_name] = original_column
                break
    return resolved


EXCEL_IMPORT_REQUIRED_COLUMNS = ("username", "full_name", "group_name")
EXCEL_IMPORT_PREVIEW_LIMIT = 12
EXCEL_IMPORT_BATCH_SIZE = 100
EXCEL_IMPORT_TOKEN_TTL_SECONDS = 30 * 60
EXCEL_IMPORT_HEADER_SCAN_LIMIT = 25
EXCEL_IMPORT_UPDATE_FIELDS = [
    "username",
    "student_id_number",
    "first_name",
    "second_name",
    "third_name",
    "full_name",
    "group_name",
    "is_student",
    "auth_is_id",
    "password",
]


def _get_excel_import_field_labels():
    return {
        "username": "username yoki talaba ID",
        "full_name": "O'quvchining F.I.SH.",
        "group_name": "akademik guruh",
    }


def _validate_excel_import_file(excel_file):
    if not excel_file:
        raise ValidationError("Excel fayl yuklanmadi.")
    if not str(excel_file.name or "").lower().endswith((".xlsx", ".xlsm")):
        raise ValidationError("Faqat .xlsx yoki .xlsm formatdagi Excel fayl yuklang.")


def _parse_excel_users_file(excel_file):
    _validate_excel_import_file(excel_file)

    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    try:
        worksheet = None
        headers = None
        header_row_index = None
        column_map = None

        for candidate_sheet in workbook.worksheets:
            for row_index, row in enumerate(candidate_sheet.iter_rows(values_only=True), start=1):
                if row_index > EXCEL_IMPORT_HEADER_SCAN_LIMIT:
                    break

                candidate_headers = [header if header is not None else "" for header in row]
                if not any(_clean_excel_value(value) for value in candidate_headers):
                    continue

                candidate_column_map = _resolve_excel_columns(candidate_headers)
                missing_columns = [
                    column for column in EXCEL_IMPORT_REQUIRED_COLUMNS if column not in candidate_column_map
                ]
                if not missing_columns:
                    worksheet = candidate_sheet
                    headers = candidate_headers
                    header_row_index = row_index
                    column_map = candidate_column_map
                    break

            if worksheet is not None:
                break

        if worksheet is None or headers is None or header_row_index is None or column_map is None:
            field_labels = _get_excel_import_field_labels()
            missing_labels = ", ".join(field_labels[column] for column in EXCEL_IMPORT_REQUIRED_COLUMNS)
            raise ValidationError(
                f"Excel ichidan kerakli ustunlar topilmadi: {missing_labels}. "
                f"Sarlavha qatori birinchi {EXCEL_IMPORT_HEADER_SCAN_LIMIT} qatorda bo'lishi kerak."
            )

        prepared_users = {}
        skipped_count = 0
        duplicate_count = 0

        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            row_data = dict(zip(headers, row))
            username = _coerce_excel_username(row_data.get(column_map["username"]))
            full_name = _clean_excel_value(row_data.get(column_map["full_name"]))
            group_name = _clean_excel_value(row_data.get(column_map["group_name"])) or None

            if not username or not full_name:
                skipped_count += 1
                continue

            first_name, second_name, third_name = _split_student_full_name(full_name)
            if username in prepared_users:
                duplicate_count += 1

            prepared_users[username] = {
                "username": username,
                "student_id_number": username,
                "first_name": first_name,
                "second_name": second_name,
                "third_name": third_name,
                "full_name": full_name,
                "group_name": group_name,
                "is_student": True,
                "auth_is_id": True,
            }

        if not prepared_users:
            raise ValidationError("Import uchun yaroqli foydalanuvchi topilmadi.")

        return {
            "headers": headers,
            "prepared_users": list(prepared_users.values()),
            "skipped_count": skipped_count,
            "duplicate_count": duplicate_count,
        }
    finally:
        workbook.close()


def _get_excel_import_preview_dir():
    directory = os.path.join(tempfile.gettempdir(), "quiz_django_excel_imports")
    os.makedirs(directory, exist_ok=True)
    return directory


def _cleanup_expired_excel_import_previews():
    now = time.time()
    directory = _get_excel_import_preview_dir()
    for file_name in os.listdir(directory):
        if not file_name.endswith(".json"):
            continue
        file_path = os.path.join(directory, file_name)
        try:
            if now - os.path.getmtime(file_path) > EXCEL_IMPORT_TOKEN_TTL_SECONDS:
                os.remove(file_path)
        except OSError:
            continue


def _get_excel_import_preview_path(token):
    if not re.fullmatch(r"[a-f0-9]{32}", token or ""):
        raise ValidationError("Import token noto'g'ri.")
    return os.path.join(_get_excel_import_preview_dir(), f"{token}.json")


def _save_excel_import_preview(payload):
    _cleanup_expired_excel_import_previews()
    token = uuid.uuid4().hex
    file_path = _get_excel_import_preview_path(token)
    with open(file_path, "w", encoding="utf-8") as preview_file:
        json.dump(payload, preview_file, ensure_ascii=False)
    return token


def _load_excel_import_preview(token):
    file_path = _get_excel_import_preview_path(token)
    if not os.path.exists(file_path):
        raise ValidationError("Import preview topilmadi yoki muddati tugagan.")

    with open(file_path, "r", encoding="utf-8") as preview_file:
        payload = json.load(preview_file)

    created_at = float(payload.get("created_at") or 0)
    if not created_at or time.time() - created_at > EXCEL_IMPORT_TOKEN_TTL_SECONDS:
        try:
            os.remove(file_path)
        except OSError:
            pass
        raise ValidationError("Import preview muddati tugagan. Excel faylni qayta tanlang.")

    return payload


def _delete_excel_import_preview(token):
    try:
        os.remove(_get_excel_import_preview_path(token))
    except (OSError, ValidationError):
        pass


def _build_excel_import_preview_payload(parsed_payload, requested_by, file_name):
    prepared_users = parsed_payload["prepared_users"]
    usernames = [item["username"] for item in prepared_users]
    existing_usernames = set(
        CustomUser.objects.filter(username__in=usernames).values_list("username", flat=True)
    )

    preview_rows = []
    for index, row in enumerate(prepared_users[:EXCEL_IMPORT_PREVIEW_LIMIT], start=1):
        is_update = row["username"] in existing_usernames
        preview_rows.append(
            {
                "index": index,
                "username": row["username"],
                "full_name": row["full_name"],
                "group_name": row["group_name"] or "-",
                "status": "update" if is_update else "create",
                "status_label": "Yangilanadi" if is_update else "Yangi qo'shiladi",
            }
        )

    return {
        "created_at": time.time(),
        "requested_by": requested_by,
        "file_name": file_name,
        "prepared_users": prepared_users,
        "summary": {
            "total_valid": len(prepared_users),
            "new_count": sum(1 for item in prepared_users if item["username"] not in existing_usernames),
            "update_count": sum(1 for item in prepared_users if item["username"] in existing_usernames),
            "skipped_count": parsed_payload["skipped_count"],
            "duplicate_count": parsed_payload["duplicate_count"],
        },
        "preview_rows": preview_rows,
    }


def _execute_excel_import(prepared_users, progress_callback=None):
    usernames = [item["username"] for item in prepared_users]
    existing_users = {
        user.username: user
        for user in CustomUser.objects.filter(username__in=usernames)
    }
    default_password_hash = make_password(DEFAULT_USER_PASSWORD)

    created_count = 0
    updated_count = 0
    processed_count = 0
    total_count = len(prepared_users)

    for start in range(0, total_count, EXCEL_IMPORT_BATCH_SIZE):
        batch = prepared_users[start:start + EXCEL_IMPORT_BATCH_SIZE]
        users_to_create = []
        users_to_update = []

        for payload in batch:
            username = payload["username"]
            if username in existing_users:
                user = existing_users[username]
                for field_name, value in payload.items():
                    setattr(user, field_name, value)
                user.password = default_password_hash
                users_to_update.append(user)
            else:
                users_to_create.append(CustomUser(password=default_password_hash, **payload))

        with transaction.atomic():
            if users_to_create:
                CustomUser.objects.bulk_create(users_to_create, batch_size=EXCEL_IMPORT_BATCH_SIZE)
            if users_to_update:
                CustomUser.objects.bulk_update(
                    users_to_update,
                    EXCEL_IMPORT_UPDATE_FIELDS,
                    batch_size=EXCEL_IMPORT_BATCH_SIZE,
                )

        created_count += len(users_to_create)
        updated_count += len(users_to_update)
        processed_count += len(batch)

        if progress_callback:
            progress_callback(processed_count, total_count, created_count, updated_count)

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "processed_count": processed_count,
    }


@method_decorator(login_required, name='dispatch')
class ImportUsersExcelPreviewView(View):
    def post(self, request):
        try:
            parsed_payload = _parse_excel_users_file(request.FILES.get("file"))
            preview_payload = _build_excel_import_preview_payload(
                parsed_payload,
                requested_by=request.user.id,
                file_name=request.FILES["file"].name,
            )
            token = _save_excel_import_preview(preview_payload)
            return JsonResponse(
                {
                    "success": True,
                    "token": token,
                    "file_name": preview_payload["file_name"],
                    "summary": preview_payload["summary"],
                    "preview_rows": preview_payload["preview_rows"],
                    "has_more_rows": preview_payload["summary"]["total_valid"] > len(preview_payload["preview_rows"]),
                    "preview_limit": EXCEL_IMPORT_PREVIEW_LIMIT,
                }
            )
        except ValidationError as error:
            return JsonResponse({"success": False, "message": str(error)}, status=400)
        except Exception as error:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Preview tayyorlashda xatolik yuz berdi: {error}",
                },
                status=500,
            )


@method_decorator(login_required, name='dispatch')
class ImportUsersExcelStreamView(View):
    def get(self, request):
        token = (request.GET.get("token") or "").strip()

        try:
            preview_payload = _load_excel_import_preview(token)
            if preview_payload.get("requested_by") != request.user.id:
                raise ValidationError("Bu import preview sizga tegishli emas.")
        except ValidationError as error:
            return JsonResponse({"success": False, "message": str(error)}, status=400)

        prepared_users = preview_payload["prepared_users"]
        summary = preview_payload["summary"]

        def stream():
            try:
                yield (
                    "data: "
                    + json.dumps(
                        {
                            "progress": 0,
                            "status": f"{preview_payload['file_name']} importga tayyorlanmoqda...",
                        },
                        ensure_ascii=False,
                    )
                    + "\n\n"
                )
                usernames = [item["username"] for item in prepared_users]
                existing_users = {
                    user.username: user
                    for user in CustomUser.objects.filter(username__in=usernames)
                }
                default_password_hash = make_password(DEFAULT_USER_PASSWORD)
                created_count = 0
                updated_count = 0
                processed_count = 0
                total_count = len(prepared_users)

                for start in range(0, total_count, EXCEL_IMPORT_BATCH_SIZE):
                    batch = prepared_users[start:start + EXCEL_IMPORT_BATCH_SIZE]
                    users_to_create = []
                    users_to_update = []

                    for payload in batch:
                        username = payload["username"]
                        if username in existing_users:
                            user = existing_users[username]
                            for field_name, value in payload.items():
                                setattr(user, field_name, value)
                            user.password = default_password_hash
                            users_to_update.append(user)
                        else:
                            users_to_create.append(CustomUser(password=default_password_hash, **payload))

                    with transaction.atomic():
                        if users_to_create:
                            CustomUser.objects.bulk_create(users_to_create, batch_size=EXCEL_IMPORT_BATCH_SIZE)
                        if users_to_update:
                            CustomUser.objects.bulk_update(
                                users_to_update,
                                EXCEL_IMPORT_UPDATE_FIELDS,
                                batch_size=EXCEL_IMPORT_BATCH_SIZE,
                            )

                    created_count += len(users_to_create)
                    updated_count += len(users_to_update)
                    processed_count += len(batch)

                    yield (
                        "data: "
                        + json.dumps(
                            {
                                "progress": round((processed_count / total_count) * 100) if total_count else 100,
                                "processed_count": processed_count,
                                "total_count": total_count,
                                "created_count": created_count,
                                "updated_count": updated_count,
                                "status": (
                                    f"{processed_count}/{total_count} ta foydalanuvchi qayta ishladi. "
                                    f"Yangi: {created_count}, yangilandi: {updated_count}."
                                ),
                            },
                            ensure_ascii=False,
                        )
                        + "\n\n"
                    )

                yield (
                    "data: "
                    + json.dumps(
                        {
                            "progress": 100,
                            "success": True,
                            "created_count": created_count,
                            "updated_count": updated_count,
                            "skipped_count": summary["skipped_count"],
                            "duplicate_count": summary["duplicate_count"],
                            "status": (
                                f"Import yakunlandi: {created_count} ta qo'shildi, "
                                f"{updated_count} ta yangilandi, "
                                f"{summary['skipped_count']} ta o'tkazib yuborildi."
                            ),
                        },
                        ensure_ascii=False,
                    )
                    + "\n\n"
                )
            except Exception as error:
                yield (
                    "data: "
                    + json.dumps(
                        {
                            "progress": 100,
                            "success": False,
                            "error": True,
                            "status": f"Importda xatolik yuz berdi: {error}",
                        },
                        ensure_ascii=False,
                    )
                    + "\n\n"
                )
            finally:
                _delete_excel_import_preview(token)

        return StreamingHttpResponse(stream(), content_type="text/event-stream")


@method_decorator(login_required, name='dispatch')
class ImportUsersExcelView(View):
    def post(self, request):
        try:
            parsed_payload = _parse_excel_users_file(request.FILES.get("file"))
            result = _execute_excel_import(parsed_payload["prepared_users"])

            return JsonResponse(
                {
                    "success": True,
                    "message": (
                        f"Import yakunlandi: {result['created_count']} ta qo'shildi, "
                        f"{result['updated_count']} ta yangilandi, "
                        f"{parsed_payload['skipped_count']} ta o'tkazib yuborildi."
                    ),
                }
            )
        except ValidationError as error:
            return JsonResponse({"success": False, "message": str(error)}, status=400)
        except Exception as error:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Importda xatolik yuz berdi: {error}",
                },
                status=500,
            )
