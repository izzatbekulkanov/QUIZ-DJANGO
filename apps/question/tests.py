import json
import tempfile
import base64
from io import BytesIO
from datetime import timedelta
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import Group, Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from apps.account.models import CustomUser
from apps.logs.models import Log
from apps.question.models import (
    Answer,
    Category,
    HelpResultPlan,
    Question,
    StudentTest,
    StudentTestAssignment,
    StudentTestQuestion,
    Test,
)
from apps.question.utils.access import build_assignment_access_token
from apps.question.utils.utils import _coerce_import_questions, parse_pasted_questions
from apps.question.view.test import _ordered_student_question_answers
from apps.question.views import _delete_excel_import_preview
from core.error_views import render_error_page


class ImportParserTests(SimpleTestCase):
    def test_parse_pasted_questions_accepts_four_character_markers(self):
        pasted_text = """
Jism koordinatasi x0=-7m bo'lgan nuqtadan x o'qi bo'ylab doimiy 6 m/s tezlik bilan harakatlana boshladi. Qancha sekunddan so'ng jismning koordinatasi 5 m bo'ladi?

====#2

====3

====8

====1

++++

Velosipedchi 12 km/soat tezlikda 4 km masofani bosib o'tib, to'xtadi va 40 daqiqa dam oldi. Shundan so'ng qolgan 8 km yo'lni 8 km/soat tezlikda bosib o'tdi. Velosipedchining butun yo'l davomidagi o'rtacha tezligini (km/soat) toping.

====#6

====8

====9

====7

++++

O'q 20 sm qalinlikdagi doskani teshib o'tadi. O'qning doskaga yetib kelishdagi tezligi 200 m/s, doskadan chiqib ketishdagi tezligi esa 100 m/s. O'qning doska ichida harakatlanishidagi tezlanishi (km/s^2 da) qancha?

====#75

====40

====43

====15
""".strip()

        questions = _coerce_import_questions(parse_pasted_questions("", pasted_text))

        self.assertEqual(len(questions), 3)
        self.assertIn("Jism koordinatasi", questions[0]["text"])
        self.assertIn("Velosipedchi", questions[1]["text"])
        self.assertIn("O'q 20 sm", questions[2]["text"])

        self.assertEqual(len(questions[0]["answers"]), 4)
        self.assertEqual(
            [answer["is_correct"] for answer in questions[0]["answers"]],
            [True, False, False, False],
        )
        self.assertIn(">2<", questions[0]["answers"][0]["text"])
        self.assertIn(">6<", questions[1]["answers"][0]["text"])
        self.assertIn(">75<", questions[2]["answers"][0]["text"])

    def test_parse_pasted_questions_accepts_split_answer_markers(self):
        pasted_text = """
++++
Wie heißt du?
====
#Ich heiße Anna.
====
Ich wohne Anna.
====
Ich komme Anna.
====
Ich bin Anna wohne.
++++
Ich komme ___ Italien.
====
#aus
====
in
====
bei
====
von
""".strip()

        questions = _coerce_import_questions(parse_pasted_questions("", pasted_text))

        self.assertEqual(len(questions), 2)
        self.assertIn("Wie heißt du?", questions[0]["text"])
        self.assertIn("Ich komme ___ Italien.", questions[1]["text"])
        self.assertEqual(len(questions[0]["answers"]), 4)
        self.assertEqual(len(questions[1]["answers"]), 4)
        self.assertTrue(questions[0]["answers"][0]["is_correct"])
        self.assertTrue(questions[1]["answers"][0]["is_correct"])
        self.assertIn("Ich heiße Anna.", questions[0]["answers"][0]["text"])
        self.assertIn(">aus<", questions[1]["answers"][0]["text"])


    def test_parse_pasted_questions_accepts_variable_marker_counts(self):
        pasted_text = """
+
Ich ___ Lehrer.
=
##bin
=
kommen
=
heißen
=
spricht
+
„goodbye“ =
==
###Auf Wiedersehen
==
Danke
""".strip()

        questions = _coerce_import_questions(parse_pasted_questions("", pasted_text))

        self.assertEqual(len(questions), 2)
        self.assertIn("Ich ___ Lehrer.", questions[0]["text"])
        self.assertIn("„goodbye“ =", questions[1]["text"])
        self.assertEqual(len(questions[0]["answers"]), 4)
        self.assertEqual(len(questions[1]["answers"]), 2)
        self.assertTrue(questions[0]["answers"][0]["is_correct"])
        self.assertTrue(questions[1]["answers"][0]["is_correct"])
        self.assertIn(">bin<", questions[0]["answers"][0]["text"])
        self.assertIn("Auf Wiedersehen", questions[1]["answers"][0]["text"])

    def test_parse_pasted_questions_handles_inline_markers_and_skips_without_correct_answer(self):
        pasted_text = """
Mukarrama Turg‘unboyeva …….. yilda Farg‘onada da dunyoga kelgan

#Yunus Rajabiy.====

G‘ani Xoliqov.====

To‘xtasin Jalilov.====

Muxtor Ashrafiy.++++

Xorazm maqomni kim yaratgan?====

#Niyozjon Xo‘ja====

Yu.Rajabiy====

Fitrat====

Furqat++++

Maqom so‘zi qanday so‘zdan olingan va qanday ma’noni anglatadi?====

#Maqom arabcha so‘z bo‘lib, o‘rin joy “makon”, “parda” ma’nolarini anglatadi.====

Maqom grekcha so‘z bo‘lib, o‘rin joy “makon”, “parda” ma’nolarini anglatadi.====

Maqom forscha so‘z bo‘lib, o‘rin joy “makon”, “parda” ma’nolarini anglatadi.====

Maqom ingiliz tilidan olingan bo‘lib, “makon”, “parda” ma’nolarini anglatadi.++++

“Dugoh xusayn” qaysi maqom yo‘nalishiga kiradi? ====

#Farg‘ona – Toshkent.====

Xorazm.====

Qashqadaryo – Surxandaryo.====

Samarqand – Buxoro.++++

Xorazm maqomlari kim tamonidan notaga olingan?

#Matniyoz Yusupov.====

Quvondiq Iskandarov.====

Komiljon Otaniyozov.====

Ortiq Otajonov.++++

Xorazmda qanday ashula turkumi mavjud?

#Suvvora.===-

Katta ashula.=====

Yalla ijrochiligi.====

Lapar.++++

Yallachilik musiqa
""".strip()

        questions = _coerce_import_questions(parse_pasted_questions("", pasted_text))

        self.assertEqual(len(questions), 6)
        self.assertIn("Mukarrama Turg‘unboyeva", questions[0]["text"])
        self.assertIn("Xorazm maqomni kim yaratgan?", questions[1]["text"])
        self.assertIn("Xorazmda qanday ashula turkumi mavjud?", questions[5]["text"])
        self.assertNotIn("====", questions[0]["text"])
        self.assertNotIn("++++", questions[0]["answers"][-1]["text"])
        self.assertNotIn("Yallachilik musiqa", " ".join(question["text"] for question in questions))
        self.assertEqual(len(questions[0]["answers"]), 4)
        self.assertEqual(len(questions[5]["answers"]), 4)
        self.assertTrue(all(any(answer["is_correct"] for answer in question["answers"]) for question in questions))
        self.assertIn("Yunus Rajabiy.", questions[0]["answers"][0]["text"])
        self.assertIn("Suvvora.", questions[5]["answers"][0]["text"])

    def test_parse_pasted_questions_prefers_plain_text_when_html_collapses_lines(self):
        pasted_text = """
+++
Wie heiГџt du?
====
#Ich heiГџe Anna.
====
Ich wohne Anna.
====
Ich komme Anna.
====
Ich bin Anna wohne.
+++
Ich komme ___ Italien.
====
#aus
====
in
====
bei
====
von
""".strip()
        pasted_html = """
<div>
  <p>++++<br>Wie heiГџt du?<br>====<br>#Ich heiГџe Anna.<br>====<br>Ich wohne Anna.<br>====<br>Ich komme Anna.<br>====<br>Ich bin Anna wohne.</p>
  <p>++++<br>Ich komme ___ Italien.<br>====<br>#aus<br>====<br>in<br>====<br>bei<br>====<br>von</p>
</div>
""".strip()

        questions = _coerce_import_questions(parse_pasted_questions(pasted_html, pasted_text))

        self.assertEqual(len(questions), 2)
        self.assertEqual(len(questions[0]["answers"]), 4)
        self.assertEqual(len(questions[1]["answers"]), 4)
        self.assertTrue(questions[0]["answers"][0]["is_correct"])
        self.assertTrue(questions[1]["answers"][0]["is_correct"])


class ErrorPageViewTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_render_error_page_returns_custom_404(self):
        response = render_error_page(self.factory.get("/missing/"), status_code=404)

        self.assertEqual(response.status_code, 404)
        self.assertIn("Sahifa topilmadi", response.content.decode())
        self.assertIn("404", response.content.decode())

    def test_render_error_page_returns_custom_500(self):
        response = render_error_page(self.factory.get("/broken/"), status_code=500)

        self.assertEqual(response.status_code, 500)
        self.assertIn("Server xatosi", response.content.decode())


class ImportQuestionsFlowTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="question-import-admin",
            password="secret123",
            is_staff=True,
        )
        self.client.force_login(self.admin_user)
        self.category = Category.objects.create(name="Import test category")
        self.test = Test.objects.create(
            category=self.category,
            name="Import test",
            created_by=self.admin_user,
        )

    def _sample_pasted_text(self):
        return """
+++
Wie heißt du?
====
#Ich heiße Anna.
====
Ich wohne Anna.
====
Ich komme Anna.
====
Ich bin Anna wohne.
+++
Ich komme ___ Italien.
====
#aus
====
in
====
bei
====
von
""".strip()

    def test_import_preview_returns_token_and_bulk_save_persists_questions(self):
        preview_response = self.client.post(
            reverse("administrator:import-questions", args=[self.test.id]),
            {
                "pasted_html": "",
                "pasted_text": self._sample_pasted_text(),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertTrue(preview_payload["success"])
        self.assertTrue(preview_payload["preview_token"])
        self.assertEqual(preview_payload["total_questions"], 2)

        save_response = self.client.post(
            reverse("administrator:save-imported-questions"),
            data=json.dumps(
                {
                    "preview_token": preview_payload["preview_token"],
                    "correct_answer_indexes": [[0], [0]],
                }
            ),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(save_response.status_code, 200)
        save_payload = save_response.json()
        self.assertTrue(save_payload["success"])
        self.assertEqual(save_payload["saved_count"], 2)
        self.assertEqual(Question.objects.filter(test=self.test).count(), 2)
        self.assertEqual(Answer.objects.filter(question__test=self.test).count(), 8)

    def test_bulk_save_skips_duplicate_questions(self):
        preview_response = self.client.post(
            reverse("administrator:import-questions", args=[self.test.id]),
            {
                "pasted_html": "",
                "pasted_text": self._sample_pasted_text(),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        preview_token = preview_response.json()["preview_token"]

        self.client.post(
            reverse("administrator:save-imported-questions"),
            data=json.dumps(
                {
                    "preview_token": preview_token,
                    "correct_answer_indexes": [[0], [0]],
                }
            ),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        second_preview_response = self.client.post(
            reverse("administrator:import-questions", args=[self.test.id]),
            {
                "pasted_html": "",
                "pasted_text": self._sample_pasted_text(),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        second_preview_token = second_preview_response.json()["preview_token"]

        second_save_response = self.client.post(
            reverse("administrator:save-imported-questions"),
            data=json.dumps(
                {
                    "preview_token": second_preview_token,
                    "correct_answer_indexes": [[0], [0]],
                }
            ),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(second_save_response.status_code, 200)
        second_save_payload = second_save_response.json()
        self.assertTrue(second_save_payload["success"])
        self.assertEqual(second_save_payload["saved_count"], 0)
        self.assertEqual(second_save_payload["skipped_duplicates"], 2)

    def test_download_office_helper_exe_returns_file_when_built(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            exe_path = Path(temp_dir) / "QuizOfficeHelper.exe"
            exe_path.write_bytes(b"MZ-test-helper")

            with patch("apps.question.utils.utils.OFFICE_HELPER_EXE_PATH", exe_path):
                response = self.client.get(reverse("administrator:download-office-helper-exe"))

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["Content-Type"], "application/octet-stream")
            self.assertIn("QuizOfficeHelper.exe", response["Content-Disposition"])
            response.close()


class QuestionDeletionApiTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="question-delete-admin",
            password="secret123",
            is_staff=True,
        )
        self.client.force_login(self.admin_user)
        self.category = Category.objects.create(name="Question delete category")
        self.test = Test.objects.create(
            category=self.category,
            name="Question delete test",
            created_by=self.admin_user,
        )
        self.question_one = self._create_question("Question 1")
        self.question_two = self._create_question("Question 2")

    def _create_question(self, text):
        question = Question.objects.create(test=self.test, text=text)
        Answer.objects.create(question=question, text="Correct answer", is_correct=True)
        Answer.objects.create(question=question, text="Wrong answer", is_correct=False)
        return question

    def _bulk_delete_url(self):
        return reverse("administrator:questions-bulk-delete", args=[self.test.id])

    def _single_delete_url(self, question):
        return reverse("administrator:question-delete", args=[self.test.id, question.id])

    def _create_student_result(self, question):
        student = CustomUser.objects.create_user(
            username="question-delete-student",
            password="secret123",
            is_student=True,
        )
        assignment = StudentTestAssignment.objects.create(
            teacher=self.admin_user,
            test=self.test,
            category=self.category,
            total_questions=1,
            start_time=timezone.now() - timedelta(hours=1),
            end_time=timezone.now() + timedelta(hours=1),
            duration=30,
        )
        student_test = StudentTest.objects.create(
            student=student,
            assignment=assignment,
            completed=True,
            duration=60,
            score=100,
        )
        StudentTestQuestion.objects.create(
            student_test=student_test,
            question=question,
            selected_answer=question.answers.filter(is_correct=True).first(),
            is_correct=True,
            order=1,
        )

    def test_bulk_delete_requires_delete_confirmation(self):
        response = self.client.post(
            self._bulk_delete_url(),
            data=json.dumps({"confirmation": "WRONG"}),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["success"])
        self.assertEqual(Question.objects.filter(test=self.test).count(), 2)

    def test_bulk_delete_removes_all_unused_questions(self):
        response = self.client.post(
            self._bulk_delete_url(),
            data=json.dumps({"confirmation": "DELETE"}),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertEqual(Question.objects.filter(test=self.test).count(), 0)
        self.assertEqual(Answer.objects.filter(question__test=self.test).count(), 0)

    def test_bulk_delete_blocks_when_student_results_exist(self):
        self._create_student_result(self.question_one)

        response = self.client.post(
            self._bulk_delete_url(),
            data=json.dumps({"confirmation": "DELETE"}),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 409)
        payload = response.json()
        self.assertFalse(payload["success"])
        self.assertEqual(payload["used_count"], 1)
        self.assertEqual(Question.objects.filter(test=self.test).count(), 2)

    def test_single_delete_blocks_when_student_results_exist(self):
        self._create_student_result(self.question_one)

        response = self.client.post(
            self._single_delete_url(self.question_one),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 409)
        payload = response.json()
        self.assertFalse(payload["success"])
        self.assertEqual(payload["used_count"], 1)
        self.assertTrue(Question.objects.filter(id=self.question_one.id).exists())


class TestQuestionsViewPaginationTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="questions-view-admin",
            password="secret123",
            is_staff=True,
        )
        self.client.force_login(self.admin_user)
        self.category = Category.objects.create(name="Questions pagination category")
        self.test = Test.objects.create(
            category=self.category,
            name="Large question bank",
            created_by=self.admin_user,
        )

        for index in range(45):
            question = Question.objects.create(
                test=self.test,
                text=f"Pagination question {index + 1}",
            )
            Answer.objects.create(question=question, text="Correct", is_correct=True)
            Answer.objects.create(question=question, text="Wrong", is_correct=False)

    def test_questions_view_paginates_large_question_bank(self):
        response = self.client.get(reverse("administrator:test-questions", args=[self.test.id]))

        self.assertEqual(response.status_code, 200)
        questions_page = response.context["questions"]
        self.assertEqual(questions_page.paginator.count, 45)
        self.assertEqual(questions_page.paginator.per_page, 40)
        self.assertEqual(len(questions_page.object_list), 40)
        self.assertEqual(response.context["question_index_offset"], 0)

    def test_questions_view_supports_search_and_custom_page_size(self):
        response = self.client.get(
            reverse("administrator:test-questions", args=[self.test.id]),
            {"search": "Pagination question 44", "page_size": "20"},
        )

        self.assertEqual(response.status_code, 200)
        questions_page = response.context["questions"]
        self.assertEqual(questions_page.paginator.count, 1)
        self.assertEqual(questions_page.paginator.per_page, 20)
        self.assertEqual(len(questions_page.object_list), 1)
        self.assertEqual(questions_page.object_list[0].text, "Pagination question 44")


class ImportUsersExcelViewTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="admin",
            password="secret123",
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_login(self.admin_user)

    def _build_excel_file(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header) for header in headers])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)
        return SimpleUploadedFile(
            "students.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_excel_file_with_offset_headers(self, rows, blank_rows=1, use_second_sheet=False):
        workbook = Workbook()
        first_sheet = workbook.active

        if use_second_sheet:
            first_sheet.title = "Bo'sh varaq"
            worksheet = workbook.create_sheet(title="Import")
        else:
            worksheet = first_sheet

        for _ in range(blank_rows):
            worksheet.append([])

        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header) for header in headers])

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)
        return SimpleUploadedFile(
            "students-offset.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_import_creates_login_ready_student_from_excel(self):
        upload = self._build_excel_file(
            [
                {
                    "username": 4440123456807,
                    "O‘quvchining F.I.SH.": "Aliyeva Gulsanam Elmurod qizi",
                    "akademik guruh": "Biologiya",
                }
            ]
        )

        response = self.client.post(
            reverse("administrator:import-users-excel"),
            {"file": upload},
        )

        self.assertEqual(response.status_code, 200)
        student = CustomUser.objects.get(username="4440123456807")
        self.assertEqual(student.student_id_number, "4440123456807")
        self.assertEqual(student.first_name, "Gulsanam")
        self.assertEqual(student.second_name, "Aliyeva")
        self.assertEqual(student.third_name, "Elmurod qizi")
        self.assertEqual(student.group_name, "Biologiya")
        self.assertTrue(student.is_student)
        self.assertTrue(student.auth_is_id)
        self.assertTrue(student.check_password("namdpi451"))

    def test_import_updates_existing_student_and_resets_default_password(self):
        student = CustomUser.objects.create_user(
            username="4440123456798",
            password="old-password",
            auth_is_id=False,
        )
        upload = self._build_excel_file(
            [
                {
                    "username": "4440123456798",
                    "O'quvchining F.I.SH.": "Abduraufova Rayxona Zohirxon qizi",
                    "akademik guruhi": "Biologiya",
                }
            ]
        )

        response = self.client.post(
            reverse("administrator:import-users-excel"),
            {"file": upload},
        )

        self.assertEqual(response.status_code, 200)
        student.refresh_from_db()
        self.assertEqual(student.first_name, "Rayxona")
        self.assertEqual(student.second_name, "Abduraufova")
        self.assertEqual(student.third_name, "Zohirxon qizi")
        self.assertEqual(student.group_name, "Biologiya")
        self.assertEqual(student.student_id_number, "4440123456798")
        self.assertTrue(student.is_student)
        self.assertTrue(student.auth_is_id)
        self.assertTrue(student.check_password("namdpi451"))

    def test_preview_endpoint_returns_summary_and_token(self):
        upload = self._build_excel_file(
            [
                {
                    "username": "4440123456701",
                    "O'quvchining F.I.SH.": "Karimov Azizbek Alisher o'g'li",
                    "akademik guruh": "Tarix",
                }
            ]
        )

        response = self.client.post(
            reverse("administrator:import-users-excel-preview"),
            {"file": upload},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["summary"]["total_valid"], 1)
        self.assertEqual(payload["summary"]["new_count"], 1)
        self.assertEqual(payload["summary"]["update_count"], 0)
        self.assertEqual(payload["preview_rows"][0]["username"], "4440123456701")
        _delete_excel_import_preview(payload["token"])

    def test_stream_endpoint_imports_users_from_preview_token(self):
        upload = self._build_excel_file(
            [
                {
                    "username": "4440123456702",
                    "O'quvchining F.I.SH.": "Usmonova Dilnoza Xamid qizi",
                    "akademik guruh": "Matematika",
                }
            ]
        )

        preview_response = self.client.post(
            reverse("administrator:import-users-excel-preview"),
            {"file": upload},
        )
        self.assertEqual(preview_response.status_code, 200)
        token = preview_response.json()["token"]

        response = self.client.get(
            reverse("administrator:import-users-excel-stream"),
            {"token": token},
        )

        self.assertEqual(response.status_code, 200)
        stream_content = b"".join(response.streaming_content).decode("utf-8")
        self.assertIn('"success": true', stream_content.lower())
        self.assertIn('"progress": 100', stream_content)

        student = CustomUser.objects.get(username="4440123456702")
        self.assertEqual(student.first_name, "Dilnoza")
        self.assertEqual(student.second_name, "Usmonova")
        self.assertEqual(student.group_name, "Matematika")
        self.assertTrue(student.check_password("namdpi451"))

    def test_import_finds_headers_after_blank_rows_and_non_active_sheet(self):
        upload = self._build_excel_file_with_offset_headers(
            [
                {
                    "username": "4440123456703",
                    "O'quvchining F.I.SH.": "Tojiboyeva Nilufar Shavkat qizi",
                    "akademik guruh": "Biologiya",
                }
            ],
            blank_rows=2,
            use_second_sheet=True,
        )

        response = self.client.post(
            reverse("administrator:import-users-excel-preview"),
            {"file": upload},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["summary"]["total_valid"], 1)
        self.assertEqual(payload["preview_rows"][0]["username"], "4440123456703")
        _delete_excel_import_preview(payload["token"])


class ResultsViewTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="results-admin",
            password="secret123",
            is_staff=True,
        )
        self.student = CustomUser.objects.create_user(
            username="student1",
            password="secret123",
            is_student=True,
            first_name="Ali",
            second_name="Valiyev",
        )
        self.client.force_login(self.admin_user)

        self.category = Category.objects.create(name="Matematika")
        self.test = Test.objects.create(
            category=self.category,
            name="Algebra testi",
            created_by=self.admin_user,
        )
        Question.objects.create(test=self.test, text="2+2=?")

        now = timezone.now()
        self.assignment = StudentTestAssignment.objects.create(
            teacher=self.admin_user,
            test=self.test,
            category=self.category,
            total_questions=1,
            start_time=now - timedelta(hours=2),
            end_time=now + timedelta(hours=2),
            duration=30,
        )

        self.completed_today = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=True,
            duration=180,
            score=88,
        )
        self.completed_today.start_time = now - timedelta(hours=1)
        self.completed_today.end_time = now
        self.completed_today.save(update_fields=["start_time", "end_time"])

        self.incomplete_today = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=False,
            duration=0,
            score=0,
        )
        self.incomplete_today.start_time = now
        self.incomplete_today.end_time = None
        self.incomplete_today.save(update_fields=["start_time", "end_time"])

        self.completed_yesterday = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=True,
            duration=240,
            score=72,
        )
        self.completed_yesterday.start_time = now - timedelta(days=1, hours=1)
        self.completed_yesterday.end_time = now - timedelta(days=1)
        self.completed_yesterday.save(update_fields=["start_time", "end_time"])

    def test_results_view_defaults_to_today_and_includes_incomplete_tests(self):
        response = self.client.get(reverse("administrator:results"))

        self.assertEqual(response.status_code, 200)
        student_test_ids = [student_test.id for student_test in response.context["student_tests"]]
        self.assertIn(self.completed_today.id, student_test_ids)
        self.assertIn(self.incomplete_today.id, student_test_ids)
        self.assertNotIn(self.completed_yesterday.id, student_test_ids)
        self.assertEqual(response.context["filters"]["time_filter"], "today")
        self.assertEqual(response.context["filters"]["status"], "all")

    def test_results_view_can_filter_by_incomplete_status_and_all_time(self):
        response = self.client.get(
            reverse("administrator:results"),
            {"status": "incomplete", "time_filter": "all"},
        )

        self.assertEqual(response.status_code, 200)
        student_tests = list(response.context["student_tests"])
        self.assertEqual(student_tests, [self.incomplete_today])


class RoleGroupManagementTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="roles-admin",
            password="secret123",
            is_staff=True,
            is_superuser=True,
        )
        self.student_user = CustomUser.objects.create_user(
            username="role-student",
            password="secret123",
            is_student=True,
            first_name="Ali",
            second_name="Karimov",
        )
        self.staff_user = CustomUser.objects.create_user(
            username="role-staff",
            password="secret123",
            is_teacher=True,
            first_name="Vali",
            second_name="Bekov",
        )
        self.client.force_login(self.admin_user)
        self.permissions = list(Permission.objects.order_by("id")[:2])

    def test_add_role_group_creates_group_with_selected_permissions(self):
        response = self.client.post(
            reverse("administrator:add-role-group"),
            {
                "name": "Nazoratchilar",
                "permissions": [permission.id for permission in self.permissions[:1]],
            },
        )

        role_group = Group.objects.get(name="Nazoratchilar")
        self.assertRedirects(
            response,
            reverse("administrator:role-group-detail", args=[role_group.id]),
        )
        self.assertEqual(role_group.permissions.count(), 1)
        self.assertEqual(role_group.permissions.first(), self.permissions[0])

    def test_role_group_detail_updates_permissions_and_users(self):
        role_group = Group.objects.create(name="Operatorlar")

        response = self.client.post(
            reverse("administrator:role-group-detail", args=[role_group.id]),
            {
                "name": "Operatorlar Plus",
                "permissions": [permission.id for permission in self.permissions],
                "users": [self.student_user.id, self.staff_user.id],
            },
        )

        self.assertRedirects(
            response,
            reverse("administrator:role-group-detail", args=[role_group.id]),
        )

        role_group.refresh_from_db()
        self.assertEqual(role_group.name, "Operatorlar Plus")
        self.assertEqual(
            set(role_group.permissions.values_list("id", flat=True)),
            {permission.id for permission in self.permissions},
        )
        self.assertEqual(
            set(role_group.user_set.values_list("id", flat=True)),
            {self.student_user.id, self.staff_user.id},
        )

    def test_role_user_search_returns_matching_username(self):
        response = self.client.get(
            reverse("administrator:role-user-search"),
            {"q": "role-stud"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["results"]), 1)
        self.assertEqual(payload["results"][0]["username"], "role-student")

    def test_create_default_role_groups_creates_recommended_groups(self):
        response = self.client.post(reverse("administrator:create-default-role-groups"))

        self.assertRedirects(response, reverse("administrator:roles"))
        self.assertTrue(Group.objects.filter(name="Super Administrator").exists())
        self.assertTrue(Group.objects.filter(name="System Administrator").exists())
        self.assertTrue(Group.objects.filter(name="Academic Operations Manager").exists())
        self.assertTrue(Group.objects.filter(name="Standard User").exists())

    def test_superuser_can_delete_role_group(self):
        role_group = Group.objects.create(name="Delete Me")

        response = self.client.post(reverse("administrator:delete-role-group", args=[role_group.id]))

        self.assertRedirects(response, reverse("administrator:roles"))
        self.assertFalse(Group.objects.filter(id=role_group.id).exists())

    def test_super_administrator_role_can_delete_role_group(self):
        self.client.logout()
        role_super_admin = Group.objects.create(name="Super Administrator")
        delegated_admin = CustomUser.objects.create_user(
            username="delegated-admin",
            password="secret123",
            is_staff=True,
        )
        delegated_admin.groups.add(role_super_admin)
        role_group = Group.objects.create(name="Delegated Delete")

        self.client.force_login(delegated_admin)
        response = self.client.post(reverse("administrator:delete-role-group", args=[role_group.id]))

        self.assertRedirects(response, reverse("administrator:roles"))
        self.assertFalse(Group.objects.filter(id=role_group.id).exists())


class EditUserAdminViewTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="edit-admin",
            password="secret123",
            is_staff=True,
            is_superuser=True,
        )
        self.target_user = CustomUser.objects.create_user(
            username="edit-target",
            password="secret123",
            is_student=True,
            first_name="Ali",
            second_name="Valiyev",
            group_name="Old Group",
        )
        self.role_group = Group.objects.create(name="Academic Operations Manager")
        self.client.force_login(self.admin_user)

    def test_edit_user_updates_group_and_role_groups(self):
        response = self.client.post(
            reverse("administrator:edit-user", args=[self.target_user.id]),
            {
                "username": "edit-target",
                "first_name": "Ali",
                "second_name": "Valiyev",
                "full_name": "Ali Valiyev",
                "phone_number": "998901234567",
                "email": "ali@example.com",
                "contact_email": "ali@example.com",
                "group_name": "Yangi Group",
                "is_student": "on",
                "auth_is_id": "on",
                "is_help": "on",
                "role_groups": [self.role_group.id],
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        self.target_user.refresh_from_db()
        self.assertEqual(self.target_user.group_name, "Yangi Group")
        self.assertTrue(self.target_user.is_help)
        self.assertEqual(
            list(self.target_user.groups.values_list("id", flat=True)),
            [self.role_group.id],
        )

    def test_force_delete_removes_user_and_all_related_records(self):
        self.target_user.is_teacher = True
        self.target_user.save(update_fields=["is_teacher"])
        self.target_user.groups.add(self.role_group)

        category = Category.objects.create(name="Majburiy delete")
        managed_test = Test.objects.create(
            category=category,
            name="Bog'liq test",
            created_by=self.target_user,
        )
        assignment = StudentTestAssignment.objects.create(
            teacher=self.target_user,
            test=managed_test,
            category=category,
            total_questions=0,
            start_time=timezone.now() - timedelta(hours=1),
            end_time=timezone.now() + timedelta(hours=1),
            duration=30,
        )
        student_test = StudentTest.objects.create(
            student=self.target_user,
            assignment=assignment,
            completed=True,
            score=86,
        )
        Log.objects.create(
            method="POST",
            path="/administrator/users/",
            status_code=200,
            user=self.target_user,
        )

        response = self.client.post(
            reverse("administrator:force-delete-user", args=[self.target_user.id]),
            {"confirmation_username": self.target_user.username},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["redirect_url"], reverse("administrator:users"))
        self.assertFalse(CustomUser.objects.filter(id=self.target_user.id).exists())
        self.assertFalse(Test.objects.filter(id=managed_test.id).exists())
        self.assertFalse(StudentTestAssignment.objects.filter(id=assignment.id).exists())
        self.assertFalse(StudentTest.objects.filter(id=student_test.id).exists())
        self.assertFalse(Log.objects.filter(user_id=self.target_user.id).exists())
        self.assertEqual(self.role_group.user_set.count(), 0)

    def test_force_delete_requires_exact_username_confirmation(self):
        response = self.client.post(
            reverse("administrator:force-delete-user", args=[self.target_user.id]),
            {"confirmation_username": "wrong-username"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertFalse(payload["success"])
        self.assertTrue(CustomUser.objects.filter(id=self.target_user.id).exists())


class HelpUsersViewTests(TestCase):
    def setUp(self):
        self.help_user = CustomUser.objects.create_user(
            username="help-operator",
            password="secret123",
            is_help=True,
            is_staff=True,
        )
        self.other_user = CustomUser.objects.create_user(
            username="ordinary-user",
            password="secret123",
        )
        self.ungrouped_student = CustomUser.objects.create_user(
            username="ungrouped-student",
            password="secret123",
            is_student=True,
            first_name="No",
            second_name="Group",
            group_name="",
        )
        self.grouped_student = CustomUser.objects.create_user(
            username="grouped-student",
            password="secret123",
            is_student=True,
            first_name="Grouped",
            second_name="Student",
            group_name="AA-101",
        )

        self.category = Category.objects.create(name="Help category")
        self.teacher = CustomUser.objects.create_user(
            username="help-teacher",
            password="secret123",
            is_teacher=True,
            is_staff=True,
        )
        self.test_one = Test.objects.create(
            category=self.category,
            name="Matematika testi",
            created_by=self.teacher,
        )
        self.test_two = Test.objects.create(
            category=self.category,
            name="Fizika testi",
            created_by=self.teacher,
        )
        for index in range(3):
            Question.objects.create(test=self.test_one, text=f"Matematika savol {index + 1}")
            Question.objects.create(test=self.test_two, text=f"Fizika savol {index + 1}")
        self.test_one.students.add(self.ungrouped_student)
        self.test_two.students.add(self.ungrouped_student)

        now = timezone.now()
        self.assignment_one = StudentTestAssignment.objects.create(
            teacher=self.teacher,
            test=self.test_one,
            category=self.category,
            total_questions=2,
            start_time=now - timedelta(days=1),
            end_time=now + timedelta(days=1),
            duration=30,
            attempts=3,
        )
        self.assignment_two = StudentTestAssignment.objects.create(
            teacher=self.teacher,
            test=self.test_two,
            category=self.category,
            total_questions=3,
            start_time=now - timedelta(days=2),
            end_time=now + timedelta(days=2),
            duration=40,
            attempts=2,
        )
        self.student_result = StudentTest.objects.create(
            student=self.ungrouped_student,
            assignment=self.assignment_one,
            completed=True,
            score=88.5,
            end_time=now,
        )

    def test_help_users_requires_help_access(self):
        self.client.force_login(self.other_user)

        response = self.client.get(reverse("administrator:help-users"))

        self.assertEqual(response.status_code, 404)
        self.assertContains(response, "Sahifa topilmadi", status_code=404)

        response = self.client.get(reverse("administrator:help-plans"))

        self.assertEqual(response.status_code, 404)

    def test_help_only_user_can_open_help_pages(self):
        help_only_user = CustomUser.objects.create_user(
            username="help-only-user",
            password="secret123",
            is_help=True,
        )
        self.client.force_login(help_only_user)

        response = self.client.get(reverse("administrator:help-users"))

        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse("administrator:help-plans"))

        self.assertEqual(response.status_code, 200)

    def test_help_users_defaults_to_ungrouped_and_can_filter_by_group(self):
        self.client.force_login(self.help_user)

        response = self.client.get(reverse("administrator:help-users"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [user.username for user in response.context["users"].object_list],
            ["help-operator", "help-teacher", "ordinary-user", "ungrouped-student"],
        )

        response = self.client.get(
            reverse("administrator:help-users"),
            {"group_name": "AA-101"},
        )

        self.assertEqual(
            [user.username for user in response.context["users"].object_list],
            ["grouped-student"],
        )

    def test_help_user_detail_shows_assignments_and_result_summary(self):
        self.client.force_login(self.help_user)

        response = self.client.get(reverse("administrator:help-user-detail", args=[self.ungrouped_student.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Matematika testi")
        self.assertContains(response, "Random savollar")
        self.assertContains(response, "help-table-badge")
        self.assertContains(response, "88.50%")
        self.assertNotContains(response, 'name="assignment_id"', html=False)
        self.assertNotContains(response, 'id="help-assignment-filter"', html=False)
        self.assertContains(
            response,
            reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id]),
        )
        self.assertContains(
            response,
            reverse("administrator:view-test-details", args=[self.student_result.id]),
        )

        response = self.client.get(
            reverse("administrator:help-user-detail", args=[self.ungrouped_student.id]),
            {"test_id": self.test_two.id},
        )

        self.assertContains(response, "Fizika testi")
        self.assertNotContains(response, reverse("administrator:view-test-details", args=[self.student_result.id]))

    def test_help_assignment_detail_renders_new_calculator_page(self):
        self.client.force_login(self.help_user)

        response = self.client.get(
            reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tez foiz hisoblagich")
        self.assertContains(response, 'id="help-total-questions"', html=False)
        self.assertContains(response, 'id="help-scored-questions"', html=False)
        self.assertContains(response, 'max="2"', html=False)
        self.assertContains(response, 'data-max-questions="2"', html=False)
        self.assertContains(response, "Yordam natija qaydlari")
        self.assertContains(response, "Saqlash")

    def test_help_assignment_detail_requires_assigned_test_for_target_user(self):
        self.client.force_login(self.help_user)

        response = self.client.get(
            reverse("administrator:help-assignment-detail", args=[self.grouped_student.id, self.assignment_one.id])
        )

        self.assertEqual(response.status_code, 404)

    def test_help_assignment_detail_saves_result_plan(self):
        self.client.force_login(self.help_user)

        response = self.client.post(
            reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id]),
            {"target_correct_answers": "1", "status": HelpResultPlan.STATUS_PENDING},
        )

        self.assertEqual(response.status_code, 302)
        plan = HelpResultPlan.objects.get(student=self.ungrouped_student, assignment=self.assignment_one)
        self.assertEqual(plan.test, self.test_one)
        self.assertEqual(plan.total_questions, 2)
        self.assertEqual(plan.target_correct_answers, 1)
        self.assertEqual(plan.status, HelpResultPlan.STATUS_PENDING)
        self.assertEqual(plan.created_by, self.help_user)

        response = self.client.get(
            reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id])
        )

        self.assertContains(response, "1 / 2 ta")
        self.assertContains(response, "50.0%")
        self.assertContains(response, "Kutilmoqda")

    def test_help_assignment_detail_rejects_target_above_random_questions(self):
        self.client.force_login(self.help_user)

        response = self.client.post(
            reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id]),
            {"target_correct_answers": "3", "status": HelpResultPlan.STATUS_PENDING},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(HelpResultPlan.objects.exists())

    def test_help_plans_lists_plan_with_final_result_details(self):
        question_one, question_two = list(self.test_one.questions.order_by("id")[:2])
        correct_answer = Answer.objects.create(question=question_one, text="To'g'ri", is_correct=True)
        wrong_answer = Answer.objects.create(question=question_two, text="Noto'g'ri", is_correct=False)
        StudentTestQuestion.objects.create(
            student_test=self.student_result,
            question=question_one,
            selected_answer=correct_answer,
            is_correct=True,
            order=1,
        )
        StudentTestQuestion.objects.create(
            student_test=self.student_result,
            question=question_two,
            selected_answer=wrong_answer,
            is_correct=False,
            order=2,
        )
        plan = HelpResultPlan.objects.create(
            student=self.ungrouped_student,
            assignment=self.assignment_one,
            test=self.test_one,
            total_questions=self.assignment_one.total_questions,
            target_correct_answers=1,
            status=HelpResultPlan.STATUS_COMPLETED,
            created_by=self.help_user,
        )
        self.client.force_login(self.help_user)

        response = self.client.get(reverse("administrator:help-plans"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Yordam qaydlari")
        self.assertContains(response, "Matematika testi")
        self.assertContains(response, "1 / 2 ta")
        self.assertContains(response, "88.50%")
        self.assertContains(response, self.help_user.username)
        self.assertContains(response, reverse("administrator:delete-help-plan", args=[plan.id]))
        self.assertContains(response, 'id="deleteHelpPlanModal"', html=False)

    def test_delete_pending_help_plan_requires_delete_confirmation_and_removes_logs(self):
        plan = HelpResultPlan.objects.create(
            student=self.ungrouped_student,
            assignment=self.assignment_one,
            test=self.test_one,
            total_questions=self.assignment_one.total_questions,
            target_correct_answers=1,
            status=HelpResultPlan.STATUS_PENDING,
            created_by=self.help_user,
        )
        detail_url = reverse("administrator:help-assignment-detail", args=[self.ungrouped_student.id, self.assignment_one.id])
        delete_url = reverse("administrator:delete-help-plan", args=[plan.id])
        Log.objects.create(method="GET", path=detail_url, status_code=200, user=self.help_user)
        Log.objects.create(method="POST", path=delete_url, status_code=302, user=self.help_user)
        self.client.force_login(self.help_user)

        response = self.client.post(delete_url, {"confirmation": "WRONG"})

        self.assertEqual(response.status_code, 302)
        self.assertTrue(HelpResultPlan.objects.filter(id=plan.id).exists())

        response = self.client.post(delete_url, {"confirmation": "DELETE"})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(HelpResultPlan.objects.filter(id=plan.id).exists())
        self.assertFalse(Log.objects.filter(path__icontains=detail_url).exists())
        self.assertFalse(Log.objects.filter(path__icontains=delete_url).exists())

    def test_delete_all_help_plans_requires_confirmation_and_removes_help_logs(self):
        HelpResultPlan.objects.create(
            student=self.ungrouped_student,
            assignment=self.assignment_one,
            test=self.test_one,
            total_questions=self.assignment_one.total_questions,
            target_correct_answers=1,
            status=HelpResultPlan.STATUS_PENDING,
            created_by=self.help_user,
        )
        HelpResultPlan.objects.create(
            student=self.ungrouped_student,
            assignment=self.assignment_two,
            test=self.test_two,
            total_questions=self.assignment_two.total_questions,
            target_correct_answers=2,
            status=HelpResultPlan.STATUS_COMPLETED,
            created_by=self.help_user,
        )
        help_url = reverse("administrator:help-users")
        Log.objects.create(method="GET", path=help_url, status_code=200, user=self.help_user)
        self.client.force_login(self.help_user)
        delete_all_url = reverse("administrator:delete-all-help-plans")

        response = self.client.post(delete_all_url, {"confirmation": "NO"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(HelpResultPlan.objects.count(), 2)

        response = self.client.post(delete_all_url, {"confirmation": "DELETE"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(HelpResultPlan.objects.count(), 0)
        self.assertFalse(Log.objects.filter(path__icontains=help_url).exists())


class ViewTestDetailsViewTests(TestCase):
    TINY_PNG = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+aF9sAAAAASUVORK5CYII="
    )

    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="detail-admin",
            password="secret123",
            is_staff=True,
        )
        self.student = CustomUser.objects.create_user(
            username="detail-student",
            password="secret123",
            is_student=True,
            full_name="Test Student",
        )
        self.client.force_login(self.admin_user)

        self.category = Category.objects.create(name="Fizika")
        self.test = Test.objects.create(
            category=self.category,
            name="Mexanika",
            created_by=self.admin_user,
        )

        question_one = Question.objects.create(test=self.test, text="1-savol")
        question_two = Question.objects.create(test=self.test, text="2-savol")
        question_three = Question.objects.create(test=self.test, text="3-savol")

        correct_one = Answer.objects.create(question=question_one, text="A", is_correct=True)
        Answer.objects.create(question=question_one, text="B", is_correct=False)
        Answer.objects.create(question=question_two, text="A", is_correct=True)
        wrong_two = Answer.objects.create(question=question_two, text="B", is_correct=False)
        Answer.objects.create(question=question_three, text="A", is_correct=True)
        Answer.objects.create(question=question_three, text="B", is_correct=False)

        now = timezone.now()
        self.assignment = StudentTestAssignment.objects.create(
            teacher=self.admin_user,
            test=self.test,
            category=self.category,
            total_questions=3,
            start_time=now - timedelta(hours=1),
            end_time=now + timedelta(hours=1),
            duration=30,
        )
        self.student_test = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=False,
            duration=95,
            score=0,
        )

        self.correct_student_question = StudentTestQuestion.objects.create(
            student_test=self.student_test,
            question=question_one,
            selected_answer=correct_one,
            is_correct=True,
            order=2,
        )
        self.incorrect_student_question = StudentTestQuestion.objects.create(
            student_test=self.student_test,
            question=question_two,
            selected_answer=wrong_two,
            is_correct=False,
            order=1,
        )
        self.unanswered_student_question = StudentTestQuestion.objects.create(
            student_test=self.student_test,
            question=question_three,
            selected_answer=None,
            is_correct=False,
            order=3,
        )

    def test_view_test_details_builds_clear_summary_context(self):
        response = self.client.get(
            reverse("administrator:view-test-details", args=[self.student_test.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_questions"], 3)
        self.assertEqual(response.context["answered_count"], 2)
        self.assertEqual(response.context["correct_answers"], 1)
        self.assertEqual(response.context["incorrect_answers"], 1)
        self.assertEqual(response.context["unanswered_count"], 1)
        self.assertEqual(response.context["progress_percent"], 66.7)
        self.assertEqual(response.context["accuracy_percent"], 50.0)
        self.assertEqual(
            [row.order for row in response.context["question_rows"]],
            [1, 2, 3],
        )

    def test_view_test_details_renders_result_style_answers_with_snapshot_preview(self):
        self.incorrect_student_question.answered_at = timezone.now()
        self.incorrect_student_question.answer_snapshot = SimpleUploadedFile(
            "snapshot.png",
            self.TINY_PNG,
            content_type="image/png",
        )
        self.incorrect_student_question.save(update_fields=["answered_at", "answer_snapshot"])

        response = self.client.get(
            reverse("administrator:view-test-details", args=[self.student_test.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '<li class="result-answer-item result-answer-item--selected-correct">',
            html=False,
            count=1,
        )
        self.assertContains(
            response,
            '<li class="result-answer-item result-answer-item--selected-wrong">',
            html=False,
            count=1,
        )
        self.assertContains(response, 'id="answerSnapshotPreview"', html=False)
        self.assertContains(response, 'data-bs-target="#answerSnapshotPreview"', html=False)
        self.assertContains(response, 'class="result-answer-time"', html=False, count=1)
        self.assertContains(response, self.incorrect_student_question.answer_snapshot.url)


class StudentResultPageViewTests(TestCase):
    def setUp(self):
        self.student = CustomUser.objects.create_user(
            username="result-page-student",
            password="secret123",
            is_student=True,
        )
        self.teacher = CustomUser.objects.create_user(
            username="result-page-teacher",
            password="secret123",
            is_staff=True,
        )
        self.category = Category.objects.create(name="Natija sahifa")
        self.test = Test.objects.create(
            category=self.category,
            name="Natija dizayn testi",
            created_by=self.teacher,
        )

        self.question_one = Question.objects.create(test=self.test, text="1-savol matni")
        self.question_two = Question.objects.create(test=self.test, text="2-savol matni")
        self.answer_one_correct = Answer.objects.create(question=self.question_one, text="To'g'ri javob", is_correct=True)
        self.answer_one_wrong = Answer.objects.create(question=self.question_one, text="Noto'g'ri javob", is_correct=False)
        self.answer_two_correct = Answer.objects.create(question=self.question_two, text="Ikkinchi to'g'ri", is_correct=True)
        self.answer_two_wrong = Answer.objects.create(question=self.question_two, text="Ikkinchi noto'g'ri", is_correct=False)

        now = timezone.now()
        self.assignment = StudentTestAssignment.objects.create(
            teacher=self.teacher,
            test=self.test,
            category=self.category,
            total_questions=2,
            start_time=now - timedelta(hours=1),
            end_time=now + timedelta(hours=1),
            duration=30,
        )
        self.student_test = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=True,
            score=50,
            duration=120,
            end_time=now,
        )
        StudentTestQuestion.objects.create(
            student_test=self.student_test,
            question=self.question_one,
            selected_answer=self.answer_one_correct,
            is_correct=True,
            order=1,
        )
        StudentTestQuestion.objects.create(
            student_test=self.student_test,
            question=self.question_two,
            selected_answer=self.answer_two_wrong,
            is_correct=False,
            order=2,
        )

    def test_result_page_requires_login(self):
        response = self.client.get(reverse("landing:view_result", args=[self.student_test.id]))

        self.assertEqual(response.status_code, 302)
        self.assertIn("/auth/login/", response["Location"])

    def test_result_page_renders_question_answer_list_states(self):
        self.client.force_login(self.student)

        response = self.client.get(reverse("landing:view_result", args=[self.student_test.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Savol 1")
        self.assertContains(response, "Savol 2")
        self.assertContains(response, '<li class="result-answer-item result-answer-item--selected-correct">', html=False, count=1)
        self.assertContains(response, '<li class="result-answer-item result-answer-item--selected-wrong">', html=False, count=1)
        self.assertContains(response, 'result-answer-status result-answer-status--correct', html=False, count=2)


class SecureTestSessionFlowTests(TestCase):
    TINY_CAMERA_SNAPSHOT = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+aF9sAAAAASUVORK5CYII="
    )

    def setUp(self):
        self.student = CustomUser.objects.create_user(
            username="secure-student",
            password="secret123",
            is_student=True,
        )
        self.outsider = CustomUser.objects.create_user(
            username="secure-outsider",
            password="secret123",
            is_student=True,
        )
        self.teacher = CustomUser.objects.create_user(
            username="secure-teacher",
            password="secret123",
            is_staff=True,
        )
        self.category = Category.objects.create(name="Secure Session")
        self.test = Test.objects.create(
            category=self.category,
            name="Secure Start Test",
            created_by=self.teacher,
        )
        self.question = Question.objects.create(test=self.test, text="Qaysi javob to'g'ri?")
        self.correct_answer = Answer.objects.create(question=self.question, text="To'g'ri", is_correct=True)
        self.wrong_answer = Answer.objects.create(question=self.question, text="Noto'g'ri", is_correct=False)
        self.assignment = StudentTestAssignment.objects.create(
            teacher=self.teacher,
            test=self.test,
            category=self.category,
            total_questions=1,
            start_time=timezone.now() - timedelta(hours=1),
            end_time=timezone.now() + timedelta(hours=1),
            duration=30,
        )
        self.test.students.add(self.student)

    def _create_help_plan_attempt(self, selected_states, target_correct_answers):
        category = Category.objects.create(name=f"Help Plan {len(selected_states)}")
        test = Test.objects.create(
            category=category,
            name=f"Help Plan Test {len(selected_states)}",
            created_by=self.teacher,
        )
        question_items = []
        correct_answers = []
        wrong_answers = []
        for index in range(1, len(selected_states) + 1):
            question = Question.objects.create(test=test, text=f"Yordam savol {index}")
            correct_answer = Answer.objects.create(question=question, text=f"To'g'ri {index}", is_correct=True)
            wrong_answer = Answer.objects.create(question=question, text=f"Noto'g'ri {index}", is_correct=False)
            question_items.append((question, correct_answer, wrong_answer))
            correct_answers.append(correct_answer)
            wrong_answers.append(wrong_answer)

        assignment = StudentTestAssignment.objects.create(
            teacher=self.teacher,
            test=test,
            category=category,
            total_questions=len(selected_states),
            start_time=timezone.now() - timedelta(hours=1),
            end_time=timezone.now() + timedelta(hours=1),
            duration=30,
        )
        test.students.add(self.student)
        student_test = StudentTest.objects.create(
            student=self.student,
            assignment=assignment,
            completed=False,
        )

        rows = []
        for index, selected_is_correct in enumerate(selected_states, start=1):
            question, correct_answer, wrong_answer = question_items[index - 1]

            if selected_is_correct is True:
                selected_answer = correct_answer
                is_correct = True
            elif selected_is_correct is False:
                selected_answer = wrong_answer
                is_correct = False
            else:
                selected_answer = None
                is_correct = False

            rows.append(
                StudentTestQuestion.objects.create(
                    student_test=student_test,
                    question=question,
                    selected_answer=selected_answer,
                    is_correct=is_correct,
                    order=index,
                )
            )

        plan = HelpResultPlan.objects.create(
            student=self.student,
            assignment=assignment,
            test=test,
            total_questions=assignment.total_questions,
            target_correct_answers=target_correct_answers,
            status=HelpResultPlan.STATUS_PENDING,
            created_by=self.teacher,
        )
        return student_test, plan, rows, correct_answers, wrong_answers

    def test_legacy_start_url_redirects_to_secure_token_url(self):
        self.client.force_login(self.student)

        response = self.client.get(reverse("landing:start-test", args=[self.assignment.id]))

        secure_token = build_assignment_access_token(self.student.id, self.assignment.id)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("landing:start-test", args=[secure_token]))

    def test_secure_start_url_requires_assignment_access(self):
        self.client.force_login(self.outsider)

        outsider_token = build_assignment_access_token(self.outsider.id, self.assignment.id)
        response = self.client.get(reverse("landing:start-test", args=[outsider_token]))

        self.assertEqual(response.status_code, 403)

    def test_start_page_renders_finish_modal_loader_and_random_answer_order(self):
        Answer.objects.create(question=self.question, text="Noto'g'ri 2", is_correct=False)
        Answer.objects.create(question=self.question, text="Noto'g'ri 3", is_correct=False)
        self.client.force_login(self.student)

        secure_token = build_assignment_access_token(self.student.id, self.assignment.id)
        response = self.client.get(reverse("landing:start-test", args=[secure_token]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="exam-submit-loader text-center py-3"', html=False)
        self.assertNotContains(response, 'class="exam-submit-overlay"', html=False)
        self.assertContains(response, "Natija tayyorlanmoqda")

        student_question = StudentTestQuestion.objects.get(
            student_test__student=self.student,
            question=self.question,
        )
        expected_answer_ids = [answer.id for answer in _ordered_student_question_answers(student_question)]
        rendered_answer_ids = [answer["id"] for answer in response.context["questions"][0]["answers"]]
        database_answer_ids = list(self.question.answers.order_by("id").values_list("id", flat=True))

        self.assertEqual(rendered_answer_ids, expected_answer_ids)
        self.assertNotEqual(rendered_answer_ids, database_answer_ids)

    def test_save_answer_saves_snapshot_and_timestamp(self):
        self.client.force_login(self.student)
        student_test = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=False,
        )
        student_question = StudentTestQuestion.objects.create(
            student_test=student_test,
            question=self.question,
            order=1,
        )

        response = self.client.post(
            reverse("landing:save-answer"),
            data=json.dumps(
                {
                    "student_question_id": student_question.id,
                    "answer_id": self.correct_answer.id,
                    "camera_snapshot": self.TINY_CAMERA_SNAPSHOT,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        student_question.refresh_from_db()
        self.assertEqual(student_question.selected_answer_id, self.correct_answer.id)
        self.assertTrue(student_question.is_correct)
        self.assertIsNotNone(student_question.answered_at)
        self.assertTrue(bool(student_question.answer_snapshot))

    def test_save_answer_accepts_multipart_snapshot_upload(self):
        self.client.force_login(self.student)
        student_test = StudentTest.objects.create(
            student=self.student,
            assignment=self.assignment,
            completed=False,
        )
        student_question = StudentTestQuestion.objects.create(
            student_test=student_test,
            question=self.question,
            order=1,
        )
        image_bytes = base64.b64decode(self.TINY_CAMERA_SNAPSHOT.split(",", 1)[1])

        response = self.client.post(
            reverse("landing:save-answer"),
            data={
                "student_question_id": student_question.id,
                "answer_id": self.correct_answer.id,
                "camera_snapshot": SimpleUploadedFile(
                    "answer-snapshot.png",
                    image_bytes,
                    content_type="image/png",
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        student_question.refresh_from_db()
        self.assertEqual(student_question.selected_answer_id, self.correct_answer.id)
        self.assertTrue(bool(student_question.answer_snapshot))

    def test_save_answer_rejects_other_students_question(self):
        other_test = StudentTest.objects.create(
            student=self.outsider,
            assignment=self.assignment,
            completed=False,
        )
        other_question = StudentTestQuestion.objects.create(
            student_test=other_test,
            question=self.question,
            order=1,
        )
        self.client.force_login(self.student)

        response = self.client.post(
            reverse("landing:save-answer"),
            data=json.dumps(
                {
                    "student_question_id": other_question.id,
                    "answer_id": self.correct_answer.id,
                    "camera_snapshot": self.TINY_CAMERA_SNAPSHOT,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()["success"])

    def test_finish_help_plan_does_not_lower_better_real_result(self):
        student_test, plan, rows, correct_answers, wrong_answers = self._create_help_plan_attempt(
            [True, True, False],
            target_correct_answers=1,
        )
        self.client.force_login(self.student)

        response = self.client.post(reverse("landing:finish-test", args=[student_test.id]))

        self.assertEqual(response.status_code, 200)
        student_test.refresh_from_db()
        plan.refresh_from_db()
        self.assertAlmostEqual(student_test.score, 66.666666, places=4)
        self.assertEqual(plan.status, HelpResultPlan.STATUS_COMPLETED)
        self.assertEqual(
            StudentTestQuestion.objects.filter(student_test=student_test, is_correct=True).count(),
            2,
        )
        rows[2].refresh_from_db()
        self.assertEqual(rows[2].selected_answer_id, wrong_answers[2].id)
        self.assertFalse(rows[2].is_correct)

    def test_finish_help_plan_randomly_fills_only_answered_questions(self):
        student_test, plan, rows, correct_answers, wrong_answers = self._create_help_plan_attempt(
            [False, False, False, False, None],
            target_correct_answers=3,
        )
        self.client.force_login(self.student)

        with patch("apps.question.view.test.random.shuffle", side_effect=lambda items: items.reverse()):
            response = self.client.post(reverse("landing:finish-test", args=[student_test.id]))

        self.assertEqual(response.status_code, 200)
        student_test.refresh_from_db()
        plan.refresh_from_db()
        self.assertEqual(plan.status, HelpResultPlan.STATUS_COMPLETED)
        self.assertEqual(student_test.student_questions.filter(is_correct=True).count(), 3)
        self.assertAlmostEqual(student_test.score, 60.0, places=4)

        refreshed_rows = list(student_test.student_questions.order_by("order"))
        self.assertEqual(refreshed_rows[0].selected_answer_id, wrong_answers[0].id)
        self.assertFalse(refreshed_rows[0].is_correct)
        self.assertEqual(refreshed_rows[1].selected_answer_id, correct_answers[1].id)
        self.assertEqual(refreshed_rows[2].selected_answer_id, correct_answers[2].id)
        self.assertEqual(refreshed_rows[3].selected_answer_id, correct_answers[3].id)
        self.assertIsNone(refreshed_rows[4].selected_answer_id)
        self.assertFalse(refreshed_rows[4].is_correct)


class AddAssignTestViewTests(TestCase):
    def setUp(self):
        self.admin_user = CustomUser.objects.create_user(
            username="assign-admin",
            password="secret123",
            is_staff=True,
        )
        self.client.force_login(self.admin_user)
        self.category = Category.objects.create(name="Kimyo")
        self.test = Test.objects.create(
            category=self.category,
            name="Organik kimyo",
            created_by=self.admin_user,
        )
        Question.objects.create(test=self.test, text="Savol 1")

    def test_add_assignment_allows_past_start_time(self):
        now = timezone.now()
        response = self.client.post(
            reverse("administrator:add-assign-test"),
            {
                "category_id": self.category.id,
                "test_id": self.test.id,
                "total_questions": 1,
                "start_time": (now - timedelta(hours=2)).strftime("%Y-%m-%d %H:%M"),
                "end_time": (now + timedelta(hours=1)).strftime("%Y-%m-%d %H:%M"),
                "duration": 30,
                "max_attempts": 3,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(StudentTestAssignment.objects.count(), 1)
        assignment = StudentTestAssignment.objects.get()
        self.assertEqual(assignment.teacher_id, self.admin_user.id)
        self.assertEqual(assignment.test_id, self.test.id)
        self.assertEqual(assignment.category_id, self.category.id)
