import io
import json
import random
from datetime import datetime

import openpyxl
from django.db.models import F

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now, make_aware
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from question.forms import TestForm, AddQuestionForm
from question.models import Test, Category, Question, Answer, StudentTestAssignment, StudentTest, StudentTestQuestion


@method_decorator(login_required, name='dispatch')
class QuestionView(View):
    template_name = 'question/views/questions.html'

    def get(self, request):
        # Filtering parameters
        category_filter = request.GET.get('category', None)  # Filter by category
        search_query = request.GET.get('search', '')  # Search by name

        # Get all tests with question count and student count
        tests_queryset = Test.objects.select_related('category').annotate(
            question_count=Count('questions'),
            student_count=Count('students')  # Talabalar sonini hisoblash
        )

        # Apply filters
        if category_filter:
            tests_queryset = tests_queryset.filter(category__id=category_filter)
        if search_query:
            tests_queryset = tests_queryset.filter(name__icontains=search_query)

        # Pagination
        paginator = Paginator(tests_queryset, 50)  # Show 50 tests per page
        page_number = request.GET.get('page')
        tests = paginator.get_page(page_number)

        # Get all categories for filter dropdown
        categories = Category.objects.all()

        context = {
            'tests': tests,  # Paginated test objects with question and student counts
            'categories': categories,  # All categories for filtering
            'filters': {
                'category': category_filter,
                'search': search_query,
            },
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class AddQuestionView(View):
    template_name = 'question/views/add-question.html'

    def get(self, request):
        # Fetch categories for the dropdown
        categories = Category.objects.all()

        context = {
            'categories': categories,  # Categories for the dropdown
        }
        return render(request, self.template_name, context)

    def post(self, request):
        # Handle form submission for adding a new test
        category_id = request.POST.get('category')
        name = request.POST.get('name')
        description = request.POST.get('description')

        try:
            category = Category.objects.get(id=category_id)
            Test.objects.create(
                category=category,
                name=name,
                description=description,
                created_by=request.user,
            )
            return JsonResponse({"success": True, "message": "Yangi test muvaffaqiyatli qo'shildi!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Xatolik yuz berdi: {str(e)}"})


@method_decorator(login_required, name='dispatch')
class DeleteTestView(View):
    def post(self, request, test_id):
        if request.user.is_superuser:  # Only admins can delete tests
            test = get_object_or_404(Test.objects.annotate( question_count=Count('questions')), id=test_id)

            # Check if there are related questions or results
            if test.question_count == 0:
                test.delete()
                return JsonResponse({"success": True, "message": "Test muvaffaqiyatli o‘chirildi!"})
            else:
                return JsonResponse({
                    "success": False,
                    "message": "Testga bog‘langan savollar yoki natijalar mavjudligi sabab o‘chirib bo‘lmaydi."
                })
        return JsonResponse({"success": False, "message": "Ruxsat etilmagan amal!"})


@method_decorator(login_required, name='dispatch')
class EditTestView(View):
    template_name = 'question/views/edit-question.html'

    def get(self, request, test_id):
        # Fetch the test object
        test = get_object_or_404(Test, id=test_id)
        form = TestForm(instance=test)
        context = {
            'form': form,
            'test': test,
        }
        return render(request, self.template_name, context)

    def post(self, request, test_id):
        # Fetch the test object
        test = get_object_or_404(Test, id=test_id)
        form = TestForm(request.POST, instance=test)
        if form.is_valid():
            form.save()
            return JsonResponse({"success": True, "message": "Test muvaffaqiyatli yangilandi!"})
        return JsonResponse({"success": False, "message": "Formada xatolik mavjud!"})


@method_decorator(login_required, name='dispatch')
class AddQuestionTestView(View):
    template_name = 'question/views/question-add-test.html'

    def get(self, request, test_id):
        test = get_object_or_404(Test, id=test_id)
        form = AddQuestionForm()

        # Ushbu testga tegishli barcha savollar va ularga tegishli javoblarni olish
        questions = test.questions.prefetch_related('answers')

        context = {
            'form': form,
            'test': test,
            'questions': questions
        }
        return render(request, self.template_name, context)

    def post(self, request, test_id):
        test = get_object_or_404(Test, id=test_id)
        print(f"Test ID: {test_id}, Test: {test}")  # Test ma'lumotini tekshirish

        form = AddQuestionForm(request.POST, request.FILES)
        print(f"POST ma'lumotlari: {request.POST}")  # POST ma'lumotlarini tekshirish
        print(f"Fayllar: {request.FILES}")  # Yuklangan fayllarni tekshirish

        if form.is_valid():
            print("Form validatsiyadan o'tdi!")  # Form validatsiyasi muvaffaqiyatli o'tganini tekshirish

            # Savolni yaratish
            question = Question.objects.create(
                test=test,
                text=form.cleaned_data['text'],
                image=form.cleaned_data.get('image')  # Rasmlar ixtiyoriy
            )
            print(f"Yaratilgan savol: {question}")  # Yaratilgan savolni tekshirish

            # Javoblar va ularning atributlarini olish
            answers = request.POST.getlist('answers[]')  # Javob matnlari
            is_correct_flags = request.POST.getlist('is_correct[]')  # To'g'ri yoki noto'g'ri bayrog'i

            print(f"Answers: {answers}")  # Kiruvchi javoblarni tekshirish
            print(f"Is_correct_flags: {is_correct_flags}")  # To'g'ri javoblar bayrog'ini tekshirish

            # Variantlarni yaratish
            for index, answer_text in enumerate(answers):
                is_correct = is_correct_flags[index].lower() == 'true'
                print(
                    f"Javob: {answer_text.strip()}, To'g'ri: {is_correct}")  # Har bir javobni va uning holatini tekshirish

                Answer.objects.create(
                    question=question,
                    text=answer_text.strip(),
                    is_correct=is_correct
                )

            print("Savol va barcha javoblar muvaffaqiyatli saqlandi!")  # Yakuniy muvaffaqiyat xabari
            return JsonResponse({"success": True, "message": "Savol va barcha javoblar muvaffaqiyatli saqlandi!"})

        print(f"Form validatsiya xatoliklari: {form.errors}")  # Form validatsiyasi xatolarini tekshirish
        return JsonResponse({"success": False, "errors": form.errors})

@method_decorator(login_required, name='dispatch')
class AssignTestView(View):
    template_name = 'question/views/assign-test.html'

    def get(self, request):
        # Filtr parametrlari
        category_id = request.GET.get('category', None)
        test_name = request.GET.get('test_name', None)
        start_time = request.GET.get('start_time', None)
        end_time = request.GET.get('end_time', None)

        # Hammasini olish yoki filtr qo‘llash
        assignments = StudentTestAssignment.objects.select_related('category', 'test', 'teacher')

        if category_id:
            assignments = assignments.filter(category_id=category_id)
        if test_name:
            assignments = assignments.filter(test__name__icontains=test_name)
        if start_time:
            assignments = assignments.filter(start_time__gte=start_time)
        if end_time:
            assignments = assignments.filter(end_time__lte=end_time)

        # Pagination
        paginator = Paginator(assignments, 50)  # Har bir sahifada 50 ta element
        page_number = request.GET.get('page', 1)
        paginated_assignments = paginator.get_page(page_number)

        # Kategoriyalarni olish
        categories = Category.objects.all()

        context = {
            'assignments': paginated_assignments,
            'categories': categories,
            'filters': {
                'category': category_id,
                'test_name': test_name,
                'start_time': start_time,
                'end_time': end_time,
            },
        }
        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class EditAssignTestView(View):
    template_name = 'question/views/edit-assign-test.html'

    def get(self, request, assignment_id):
        # Fetch the assignment or return 404
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)
        categories = Category.objects.all()

        # Prepare context with assignment data and categories
        context = {
            'assignment': assignment,
            'categories': categories,
            'filters': {
                'category_id': str(assignment.category.id),
                'test_id': str(assignment.test.id),
                'total_questions': assignment.total_questions,
                'start_time': assignment.start_time.strftime('%Y-%m-%d %H:%M'),
                'end_time': assignment.end_time.strftime('%Y-%m-%d %H:%M'),
                'duration': assignment.duration,
                'max_attempts': assignment.attempts,
            }
        }
        return render(request, self.template_name, context)

    def post(self, request, assignment_id):
        # Fetch the assignment
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        try:
            # Parse form data
            data = request.POST
            category_id = data.get('category_id')
            test_id = data.get('test_id')
            total_questions = int(data.get('total_questions', 0))
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            duration = int(data.get('duration', 0))
            max_attempts = int(data.get('max_attempts', 0))

            # Validate inputs
            if not all([category_id, test_id, total_questions, start_time, end_time, duration, max_attempts]):
                return JsonResponse({'success': False, 'message': 'Barcha maydonlar to‘ldirilishi kerak.'})

            # Validate category and test
            category = get_object_or_404(Category, id=category_id)
            test = get_object_or_404(Test, id=test_id)

            # Validate total questions against test question count
            test_question_count = test.questions.count()  # Assuming Test model has a related 'questions' field
            if total_questions > test_question_count:
                return JsonResponse({
                    'success': False,
                    'message': f"Savollar soni testdagi savollar sonidan ({test_question_count}) oshmasligi kerak."
                })

            # Validate dates
            from datetime import datetime
            start_dt = datetime.strptime(start_time, '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(end_time, '%Y-%m-%d %H:%M')
            now = datetime.now()

            if start_dt >= end_dt:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti tugash vaqtidan keyin bo‘lishi mumkin emas.'
                })
            if end_dt <= now:
                return JsonResponse({
                    'success': False,
                    'message': 'Tugash vaqti hozirgi vaqtdan oldin bo‘lishi mumkin emas.'
                })

            # Update assignment
            assignment.category = category
            assignment.test = test
            assignment.total_questions = total_questions
            assignment.start_time = start_dt
            assignment.end_time = end_dt
            assignment.duration = duration
            assignment.max_attempts = max_attempts
            assignment.save()

            return JsonResponse({
                'success': True,
                'message': 'Topshiriq muvaffaqiyatli yangilandi.'
            })

        except (ValueError, ValidationError) as e:
            return JsonResponse({
                'success': False,
                'message': str(e) if str(e) else 'Noto‘g‘ri ma’lumotlar kiritildi.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': 'Serverda xatolik yuz berdi. Iltimos, qayta urinib ko‘ring.'
            })


@method_decorator(login_required, name='dispatch')
class ViewAssignTestView(View):
    template_name = 'question/views/view-assign-test.html'

    def get(self, request, assignment_id):
        # Fetch the assignment with related objects
        assignment = get_object_or_404(StudentTestAssignment.objects.select_related('category', 'test', 'teacher'), id=assignment_id)

        # Get filter parameters
        group_name = request.GET.get('group_name', '')
        username = request.GET.get('username', '')
        first_name = request.GET.get('first_name', '')
        full_name = request.GET.get('full_name', '')

        # Filter student tests
        student_tests = assignment.student_tests.select_related('student')
        if group_name:
            student_tests = student_tests.filter(student__group_name__icontains=group_name)
        if username:
            student_tests = student_tests.filter(student__username__icontains=username)
        if first_name:
            student_tests = student_tests.filter(student__first_name__icontains=first_name)
        if full_name:
            student_tests = student_tests.filter(student__full_name__icontains=full_name)

        # Prepare context
        context = {
            'assignment': assignment,
            'student_tests': student_tests,
            'filters': {
                'group_name': group_name,
                'username': username,
                'first_name': first_name,
                'full_name': full_name,
            }
        }
        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class ExportAssignTestView(View):
    def get(self, request, assignment_id):
        # Fetch the assignment
        assignment = get_object_or_404(StudentTestAssignment.objects.select_related('test'), id=assignment_id)

        # Apply filters
        group_name = request.GET.get('group_name', '')
        username = request.GET.get('username', '')
        first_name = request.GET.get('first_name', '')
        full_name = request.GET.get('full_name', '')

        student_tests = assignment.student_tests.select_related('student')
        if group_name:
            student_tests = student_tests.filter(student__group_name__icontains=group_name)
        if username:
            student_tests = student_tests.filter(student__username__icontains=username)
        if first_name:
            student_tests = student_tests.filter(student__first_name__icontains=first_name)
        if full_name:
            student_tests = student_tests.filter(student__full_name__icontains=full_name)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Test_{assignment.test.name}_Results"

        # Define headers
        headers = ['#', 'Username', 'Full Name', 'Group Name', 'Score', 'Completed', 'Duration (MM:SS)']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header
            ws[f"{get_column_letter(col_num)}1"].font = openpyxl.styles.Font(bold=True)

        # Populate data
        for row_num, student_test in enumerate(student_tests, 2):
            ws[f"A{row_num}"] = row_num - 1
            ws[f"B{row_num}"] = student_test.student.username
            ws[f"C{row_num}"] = student_test.student.full_name or ''
            ws[f"D{row_num}"] = student_test.student.group_name or ''
            ws[f"E{row_num}"] = student_test.score
            ws[f"F{row_num}"] = 'Yes' if student_test.completed else 'No'
            ws[f"G{row_num}"] = student_test.get_duration_display()

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Test_{assignment.test.name}_Results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        response.write(buffer.getvalue())
        return response

@method_decorator([login_required, require_POST], name='dispatch')
class DeleteAssignmentView(View):
    def post(self, request, assignment_id):
        try:
            assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

            # Agar topshiriqqa bog'liq StudentTest yozuvlari mavjud bo'lsa, o'chirishni taqiqlash
            if assignment.student_tests.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Bu topshiriqni o‘chirish mumkin emas, chunki unga tegishli talaba testlari mavjud.'
                }, status=400)

            assignment.delete()
            return JsonResponse({
                'success': True,
                'message': 'Topshiriq muvaffaqiyatli o‘chirildi!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Xatolik yuz berdi: {str(e)}'
            }, status=500)


@method_decorator(login_required, name='dispatch')
class AddAssignTestView(View):
    template_name = 'question/views/add-assign-test.html'

    def get(self, request):
        categories = Category.objects.all()
        tests = Test.objects.all()
        context = {
            'categories': categories,
            'tests': tests,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        try:
            category_id = request.POST.get('category_id')
            test_id = request.POST.get('test_id')
            total_questions = int(request.POST.get('total_questions', 0))
            start_time_str = request.POST.get('start_time')
            end_time_str = request.POST.get('end_time')
            duration = request.POST.get('duration')
            max_attempts = int(request.POST.get('max_attempts', 3))  # Yangi maydon

            category = get_object_or_404(Category, id=category_id)
            test = get_object_or_404(Test, id=test_id)

            # Savollar sonini tekshirish
            available_questions = test.questions.count()
            if total_questions > available_questions:
                return JsonResponse({
                    'success': False,
                    'message': f"Testda faqat {available_questions} ta savol mavjud."
                }, status=400)

            # Maksimal urinishlar sonini tekshirish
            if max_attempts < 1:
                return JsonResponse({
                    'success': False,
                    'message': "Maksimal urinishlar soni 1 dan kam bo‘lishi mumkin emas."
                }, status=400)

            # Vaqtni parsing qilish
            try:
                start_time = make_aware(datetime.strptime(start_time_str, '%Y-%m-%d %H:%M'))
                end_time = make_aware(datetime.strptime(end_time_str, '%Y-%m-%d %H:%M'))
            except ValueError as e:
                return JsonResponse({
                    'success': False,
                    'message': 'Vaqt formati noto‘g‘ri (YYYY-MM-DD HH:mm).'
                }, status=400)

            # Vaqt validatsiyasi
            now = timezone.now()
            if start_time < now:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti hozirgi vaqtdan keyin bo‘lishi kerak.'
                }, status=400)
            if start_time >= end_time:
                return JsonResponse({
                    'success': False,
                    'message': 'Boshlanish vaqti tugash vaqtidan oldin bo‘lishi kerak.'
                }, status=400)

            # Topshiriq allaqachon mavjudligini tekshirish
            if StudentTestAssignment.objects.filter(
                category=category,
                test=test,
                start_time=start_time,
                end_time=end_time,
            ).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Bunday test topshirig‘i allaqachon mavjud!'
                }, status=400)

            # Topshiriqni yaratish
            assignment = StudentTestAssignment.objects.create(
                teacher=request.user,
                test=test,
                category=category,
                total_questions=total_questions,
                start_time=start_time,
                end_time=end_time,
                duration=duration,
                is_active=True,
                status='pending',
                attempts=max_attempts  # Yangi maydonni saqlash
            )

            return JsonResponse({
                'success': True,
                'message': 'Test topshirig‘i muvaffaqiyatli qo‘shildi!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Xatolik yuz berdi: {str(e)}'
            }, status=500)
@method_decorator(csrf_exempt, name='dispatch')
class ToggleActiveView(View):
    def post(self, request, assignment_id):
        try:
            data = json.loads(request.body)
            is_active = data.get('is_active', None)

            if is_active is None:
                return JsonResponse({"success": False, "message": "Invalid data received."}, status=400)

            assignment = StudentTestAssignment.objects.get(id=assignment_id)
            assignment.is_active = is_active
            assignment.save()

            status = "activated" if is_active else "deactivated"
            return JsonResponse({"success": True, "message": f"The assignment was successfully {status}."})
        except StudentTestAssignment.DoesNotExist:
            return JsonResponse({"success": False, "message": "Assignment not found."}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)


@method_decorator(login_required, name='dispatch')
class StartTestDetailView(View):
    template_name = 'test/start_test_detail.html'

    def get(self, request, assignment_id):
        # StudentTestAssignment obyektini olish
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        # Tegishli test, kategoriya va o‘qituvchi maʼlumotlarini olish
        assignment_data = {
            'id': assignment.id,
            'test_name': assignment.test.name,
            'category_name': assignment.category.name,
            'teacher_name': assignment.teacher.full_name,
            'total_questions': assignment.total_questions,
            'start_time': assignment.start_time,
            'end_time': assignment.end_time,
            'duration': assignment.duration,
            'status': assignment.status,
            'is_active': assignment.is_active,
            'created_at': assignment.created_at,
        }

        # Render qilish
        return render(request, self.template_name, {'assignment': assignment_data})

@method_decorator(login_required, name='dispatch')
class StartTestView(View):
    template_name = 'test/start_test.html'

    def get(self, request, assignment_id):
        # StudentTestAssignment ni olish
        assignment = get_object_or_404(StudentTestAssignment, id=assignment_id)

        # Foydalanuvchiga tegishli yakunlanmagan testni qidirish
        unfinished_test = StudentTest.objects.filter(
            student=request.user,
            assignment=assignment,
            completed=False
        ).first()

        if unfinished_test:
            # Mavjud yakunlanmagan testni qayta ochish
            elapsed_time = (now() - unfinished_test.start_time).seconds
            remaining_time = max(0, assignment.duration * 60 - elapsed_time)  # Sekundlarda

            questions_data = []
            for student_question in unfinished_test.student_questions.order_by('order').all():
                answers = list(student_question.question.answers.all())
                questions_data.append({
                    'id': student_question.id,
                    'text': student_question.question.text,
                    'image': student_question.question.image.url if student_question.question.image else None,
                    'answers': [
                        {
                            'id': answer.id,
                            'text': answer.text,
                            'selected': student_question.selected_answer == answer  # Tanlanganmi yoki yo'q
                        }
                        for answer in answers
                    ]
                })

            context = {
                'assignment': assignment,
                'student_test': unfinished_test,
                'questions': questions_data,
                'remaining_time': remaining_time,  # Qolgan vaqtni yuborish
            }
        else:
            # Yangi test yaratish
            student_test = StudentTest.objects.create(
                student=request.user,
                assignment=assignment,
                start_time=now()
            )

            # Savollarni generatsiya qilish va tartibini saqlash
            available_questions = list(assignment.test.questions.all())
            if len(available_questions) < assignment.total_questions:
                return JsonResponse({'success': False, 'message': 'Yetarli savollar mavjud emas.'}, status=400)

            selected_questions = random.sample(available_questions, assignment.total_questions)
            for index, question in enumerate(selected_questions, start=1):
                StudentTestQuestion.objects.create(
                    student_test=student_test,
                    question=question,
                    order=index  # Tartibini saqlash
                )

            questions_data = []
            for student_question in student_test.student_questions.order_by('order').all():
                answers = list(student_question.question.answers.all())
                questions_data.append({
                    'id': student_question.id,
                    'text': student_question.question.text,
                    'image': student_question.question.image.url if student_question.question.image else None,
                    'answers': [
                        {
                            'id': answer.id,
                            'text': answer.text,
                            'selected': False  # Yangi test uchun tanlanmagan
                        }
                        for answer in answers
                    ]
                })

            context = {
                'assignment': assignment,
                'student_test': student_test,
                'questions': questions_data,
                'remaining_time': assignment.duration * 60,  # To'liq vaqt
            }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class GetRemainingTimeView(View):
    def get(self, request, test_id):
        student_test = get_object_or_404(StudentTest, id=test_id, student=request.user)

        if student_test.completed:
            return JsonResponse({'success': False, 'message': 'Test yakunlangan.'}, status=400)

        assignment = student_test.assignment
        elapsed_time = (now() - student_test.start_time).seconds
        remaining_time = max(0, assignment.duration * 60 - elapsed_time)

        return JsonResponse({'success': True, 'remaining_time': remaining_time})
@method_decorator(login_required, name='dispatch')
class SaveAnswerView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            student_question_id = data.get('student_question_id')
            answer_id = data.get('answer_id')

            # StudentTestQuestion obyektini olish
            student_question = get_object_or_404(StudentTestQuestion, id=student_question_id)
            selected_answer = get_object_or_404(Answer, id=answer_id)

            # Javobni saqlash
            student_question.selected_answer = selected_answer
            student_question.is_correct = selected_answer.is_correct
            student_question.save()

            return JsonResponse({'success': True, 'message': 'Javob muvaffaqiyatli saqlandi!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

@method_decorator(login_required, name='dispatch')
class FinishTestView(View):
    def post(self, request, test_id):
        student_test = get_object_or_404(StudentTest, id=test_id, student=request.user)

        # Testni yakunlash
        correct_answers = student_test.student_questions.filter(is_correct=True).count()
        total_questions = student_test.student_questions.count()
        score = (correct_answers / total_questions * 100) if total_questions > 0 else 0.0

        student_test.completed = True
        student_test.end_time = timezone.now()
        student_test.duration = (student_test.end_time - student_test.start_time).seconds  # Soniyalarda
        student_test.score = score
        student_test.save()

        # Natija sahifasiga yo'naltirish uchun natija URLini qaytarish
        result_url = f"/results/{student_test.id}/"
        return JsonResponse({
            'success': True,
            'message': f'Test tugatildi. Natijangiz: {score:.2f}%',
            'redirect_url': result_url
        })

@method_decorator(login_required, name='dispatch')
class TestResultView(View):
    template_name = 'test/test_result.html'

    def get(self, request, test_id):
        student_test = get_object_or_404(StudentTest, id=test_id, student=request.user)
        correct_answers = student_test.student_questions.filter(is_correct=True).count()
        total_questions = student_test.student_questions.count()

        context = {
            'student_test': student_test,
            'correct_answers': correct_answers,
            'total_questions': total_questions,
        }
        return render(request, self.template_name, context)